# !/usr/bin/env powerscript
# -*- python -*- coding: UTF-8 -*-
#
# Copyright (C) 2014 CONTACT Software GmbH
# All rights reserved.
# http://www.contact.de/
import logging
import os
import sys

from cdb import ElementsError
from cdb import ue
from cs.classification import api as classification_api
from cs.requirements.exceptions import MissingVariableValueError
from cs.requirements.richtext import RichTextVariables

from .utils import RequirementsTestCase

LOG = logging.getLogger(__name__)


class TestReports(RequirementsTestCase):
    need_classification_core = True
    def __init__(self, *args, **kwargs):
        self.process_list_before = self._get_process_list()
        super(TestReports, self).__init__(*args, **kwargs)

    def _get_process_list(self):
        import psutil
        proclist = "Process List:\n"
        for proc in psutil.process_iter():
            proclist += "PID:{pid}\tNAME: {name}\t\tSTATUS: {status}\n".format(**proc.as_dict())
        return proclist

    def log_process_list(self):
        LOG.error('Process List before all Report Tests:\n%s', self.process_list_before)
        LOG.error(self._get_process_list())

    def _generate_report(self, *args, **kwargs):
        from cs.tools.powerreports import WithPowerReports
        try:
            res = WithPowerReports.generate_report(
                report_action="cdbxml_report_download",
                *args, **kwargs
            )
            return res
        except BaseException as e:
            self.log_process_list()
            LOG.exception(e)
            raise

    def test_requirement_specification_overview_report_with_variables_without_values(self):
        """ (R00002910) Report: Requirement Specification Overview with variables without values"""
        from cs.requirements import RQMSpecification

        spec = RQMSpecification.KeywordQuery(name=u"report-test-specification")[0]
        req = spec.Requirements[0]
        variable_id = 'RQM_TEST_VARIABLE_001'
        req.SetText("cdbrqm_spec_object_desc_de", """<xhtml:div>{}</xhtml:div>""".format(
            RichTextVariables.get_variable_xhtml(variable_id=variable_id)
        ))
        with self.assertRaises(ue.Exception) as cm:
            self._generate_report(
                {
                    u"name": u"RequirementsOverview",
                    u"report_title": u"Anforderungsübersicht",
                    u"iso_code": u"de"
                },
                objects=[spec],
            )
        self.assertIn(variable_id, str(cm.exception))
        self.assertIn(str(MissingVariableValueError(variable_id=variable_id)), str(cm.exception))

    def test_requirement_specification_overview_report_with_variables_and_values(self):
        """ (R00002910) Report: Requirement Specification Overview with variables and values"""
        from cs.requirements import RQMSpecification
        spec = RQMSpecification.KeywordQuery(name=u"report-test-specification")[0]
        variable_id = 'RQM_RATING_RQM_COMMENT_EXTERN'
        req = spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0]
        req.SetText("cdbrqm_spec_object_desc_de", """<xhtml:div>{}</xhtml:div>""".format(
            RichTextVariables.get_variable_xhtml(variable_id=variable_id)
        ))
        classification_data = classification_api.get_new_classification(
            ["RQM_RATING"], narrowed=False
        )
        variable_value = '### test comment ###'
        classification_data['properties']['RQM_RATING_RQM_COMMENT_EXTERN'][0]['value'] = variable_value
        classification_api.update_classification(req, classification_data)
        res = self._generate_report(
            {
                u"name": u"RequirementsOverview",
                u"report_title": u"Anforderungsübersicht",
                u"iso_code": u"de"
            },
            objects=[spec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"D3": spec.name,
            u"D27": spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0].specobject_id,
            u"D28": spec.Requirements.KeywordQuery(specobject_id=u'RT000000003')[0].specobject_id,
            u"D29": spec.Requirements.KeywordQuery(specobject_id=u'RT000000005')[0].specobject_id,
            u"D30": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000002')[0].targetvalue_id,
            u"D31": spec.Requirements.KeywordQuery(specobject_id=u'RT000000001')[0].specobject_id,
            u"D32": spec.Requirements.KeywordQuery(specobject_id=u'RT000000004')[0].specobject_id,
            u"D33": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000001')[0].targetvalue_id,
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)
        self.assertIn(variable_value, tabelle1[u"C27"].value)

    def test_requirement_specification_overview_report(self):
        """ (R00002910) Report: Requirement Specification Overview"""
        from cs.requirements import RQMSpecification

        spec = RQMSpecification.KeywordQuery(name=u"report-test-specification")[0]
        res = self._generate_report(
            {
                u"name": u"RequirementsOverview",
                u"report_title": u"Anforderungsübersicht",
                u"iso_code": u"de"
            },
            objects=[spec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"D3": spec.name,
            u"D27": spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0].specobject_id,
            u"D28": spec.Requirements.KeywordQuery(specobject_id=u'RT000000003')[0].specobject_id,
            u"D29": spec.Requirements.KeywordQuery(specobject_id=u'RT000000005')[0].specobject_id,
            u"D30": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000002')[0].targetvalue_id,
            u"D31": spec.Requirements.KeywordQuery(specobject_id=u'RT000000001')[0].specobject_id,
            u"D32": spec.Requirements.KeywordQuery(specobject_id=u'RT000000004')[0].specobject_id,
            u"D33": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000001')[0].targetvalue_id,
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)

    def test_requirement_structure_overview_report(self):
        """ (R00002910) Report: Requirement Structure Overview"""
        from cs.requirements import RQMSpecification
        spec = RQMSpecification.KeywordQuery(name=u"report-test-specification")[0]
        rec = spec.Requirements.KeywordQuery(specobject_id=u"RT000000000")[0]
        res = self._generate_report(
            {
                u"name": u"RequirementStructureOverview",
                u"report_title": u"Anforderungsstruktur",
                u"iso_code": u"de"
            },
            objects=[rec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"D5": rec.specobject_id,
            u"D26": spec.Requirements.KeywordQuery(specobject_id=u'RT000000003')[0].specobject_id,
            u"D27": spec.Requirements.KeywordQuery(specobject_id=u'RT000000005')[0].specobject_id,
            u"D28": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000002')[0].targetvalue_id,
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)

    def test_missing_requirement_links_report(self):
        """ (R00002910) Report: Missing Requirement Links"""
        from cs.requirements import RQMSpecification
        spec = RQMSpecification.KeywordQuery(name=u"RQM ReqIF Interface (TEST)")[0]
        dlg_args = {
                u"object_names": u"Anforderung",
                u"links": u"Satisfies",
                u"types": u"cdbrqm_spec_object",
                u"specification_object_id": "",
                u"missing_rep_parent_object_id": "",
                u"category": "",
                u"weight": "",
                u"priority": "",
                u"discipline": ""
        }
        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"D6": spec.GetDescription(),
            u"D12": spec.Requirements.KeywordQuery(specobject_id=u'RT000000006')[0].specobject_id,
            u"D13": spec.Requirements.KeywordQuery(specobject_id=u'RT000000009')[0].specobject_id,
            u"D14": spec.Requirements.KeywordQuery(specobject_id=u'RT000000011')[0].specobject_id,
            u"D15": spec.Requirements.KeywordQuery(specobject_id=u'RT000000007')[0].specobject_id,
            u"D16": spec.Requirements.KeywordQuery(specobject_id=u'RT000000012')[0].specobject_id,
            u"D17": spec.Requirements.KeywordQuery(specobject_id=u'RT000000013')[0].specobject_id,
            u"D18": spec.Requirements.KeywordQuery(specobject_id=u'RT000000008')[0].specobject_id,
            u"D19": spec.Requirements.KeywordQuery(specobject_id=u'RT000000014')[0].specobject_id,
            u"D20": spec.Requirements.KeywordQuery(specobject_id=u'RT000000016')[0].specobject_id,
            u"D21": spec.Requirements.KeywordQuery(specobject_id=u'RT000000015')[0].specobject_id,
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)

        from cs.tools.semanticlinks import SemanticLink, SemanticLinkType
        source_classname = dest_classname = "cdbrqm_spec_object"
        satisfies_type = SemanticLinkType.ByKeys(subject_object_classname=source_classname, object_object_classname=dest_classname, name="Satisfies")

        SemanticLink.CreateNoResult(
            link_type_object_id=satisfies_type.cdb_object_id,
            subject_object_id=spec.Requirements.KeywordQuery(specobject_id=u'RT000000006')[0].cdb_object_id,
            object_object_id=spec.Requirements.KeywordQuery(specobject_id=u'RT000000009')[0].cdb_object_id,
            subject_object_classname=dest_classname,
            object_object_classname=source_classname
        )

        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        for k, val in cell_values_to_check.items():
            self.assertNotEqual(tabelle1[k].value, spec.Requirements.KeywordQuery(specobject_id=u'RT000000006')[0].specobject_id)

        verifies_type = SemanticLinkType.ByKeys(subject_object_classname=source_classname,
                                                object_object_classname=dest_classname, name="Verifies")
        SemanticLink.CreateNoResult(
            link_type_object_id=verifies_type.cdb_object_id,
            subject_object_id=spec.Requirements.KeywordQuery(specobject_id=u'RT000000014')[0].cdb_object_id,
            object_object_id=spec.Requirements.KeywordQuery(specobject_id=u'RT000000015')[0].cdb_object_id,
            subject_object_classname=dest_classname,
            object_object_classname=source_classname
        )

        dlg_args = {
            u"object_names": u"Anforderung",
            u"links": u"Verifies",
            u"types": u"cdbrqm_spec_object",
            u"specification_object_id": "",
            u"missing_rep_parent_object_id": "",
            u"category": "",
            u"weight": "",
            u"priority": "",
            u"discipline": ""
        }

        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        for k, val in cell_values_to_check.items():
            self.assertNotEqual(tabelle1[k].value,
                                spec.Requirements.KeywordQuery(specobject_id=u'RT000000014')[0].specobject_id)

    def test_missing_requirement_links_report_with_filters(self):
        from openpyxl import load_workbook
        from cs.requirements import RQMSpecification, RQMSpecObject
        spec = RQMSpecification.KeywordQuery(name=u"report-test-specification")[0]

        req_with_category_chapter = spec.Requirements[0]
        req_with_category_chapter.category = "Chapter"

        req_with_weight_2 = spec.Requirements[1]
        req_with_weight_2.weight = 2

        req_with_priority_could = spec.Requirements[2]
        req_with_priority_could.priority = "Could"

        req_with_discipline_engineering = spec.Requirements[3]
        req_with_discipline_engineering.discipline = "Systems Engineering"

        dlg_args = {
            u"object_names": u"Anforderung",
            u"links": u"Verifies",
            u"types": u"cdbrqm_spec_object",
            u"specification_object_id": "",
            u"missing_rep_parent_object_id": "",
            u"missing_rep_category": "",
            u"missing_rep_weight": "",
            u"missing_rep_priority": "",
            u"missing_rep_discipline": ""
        }

        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )

        cell_values_to_check = {
            u"D6": spec.GetDescription(),
            u"D12": spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0].specobject_id,
            u"D13": spec.Requirements.KeywordQuery(specobject_id=u'RT000000003')[0].specobject_id,
            u"D14": spec.Requirements.KeywordQuery(specobject_id=u'RT000000005')[0].specobject_id,
            u"D15": spec.Requirements.KeywordQuery(specobject_id=u'RT000000001')[0].specobject_id,
            u"D16": spec.Requirements.KeywordQuery(specobject_id=u'RT000000004')[0].specobject_id,
        }
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)

        dlg_args["missing_rep_parent_object_id"] = RQMSpecObject.KeywordQuery(specobject_id="RT000000000")[0].cdb_object_id
        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        parent_req_desc = RQMSpecObject.KeywordQuery(specobject_id="RT000000000")[0].GetDescription()
        self.assertEqual(parent_req_desc, tabelle1[u"D7"].value)
        self.assertEqual(u'RT000000003', tabelle1[u"D12"].value)
        self.assertEqual(u'RT000000005', tabelle1[u"D13"].value)
        self.assertEqual(None, tabelle1[u"D14"].value)

        dlg_args["missing_rep_parent_object_id"] = ""
        dlg_args["mapped_category_de"] = "Kapitel"
        dlg_args["missing_rep_category"] = "Chapter"
        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        self.assertEqual("Kapitel", tabelle1[u"D8"].value)
        self.assertEqual(req_with_category_chapter.specobject_id, tabelle1[u"D12"].value)
        self.assertEqual(None, tabelle1[u"D13"].value)

        dlg_args["mapped_category_de"] = ""
        dlg_args["missing_rep_category"] = ""
        dlg_args["missing_rep_weight"] = "2"
        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        self.assertEqual("2", tabelle1[u"D9"].value)
        self.assertEqual(req_with_weight_2.specobject_id, tabelle1[u"D12"].value)
        self.assertEqual(None, tabelle1[u"D13"].value)

        dlg_args["missing_rep_weight"] = ""
        dlg_args["missing_rep_priority"] = "Could"
        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        self.assertEqual("Could", tabelle1[u"G8"].value)
        self.assertEqual(req_with_priority_could.specobject_id, tabelle1[u"D12"].value)
        self.assertEqual(None, tabelle1[u"D13"].value)

        dlg_args["missing_rep_priority"] = ""
        dlg_args["mapped_discipline_de"] = "Systems Engineering"
        dlg_args["missing_rep_discipline"] = "Systems Engineering"
        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[spec],
            dlg_args=dlg_args
        )
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        self.assertEqual("Systems Engineering", tabelle1[u"G9"].value)
        self.assertEqual(req_with_discipline_engineering.specobject_id, tabelle1[u"D12"].value)
        self.assertEqual(None, tabelle1[u"D13"].value)

    def test_missing_requirement_links_overview_report(self):
        """ (R00002910) Report: Missing Requirement Links"""
        from cs.requirements import RQMSpecification
        spec = RQMSpecification.KeywordQuery(name=u"RQM ReqIF Interface (TEST)")[0]
        req = spec.Requirements.KeywordQuery(specobject_id=u'RT000000006')[0]
        res = self._generate_report(
            {
                u"name": u"MissingRequirementLinks",
                u"report_title": u"Fehlende Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[req],
            dlg_args={
                u"object_names": u"Anforderung",
                u"links": u"Satisfies",
                u"types": u"cdbrqm_spec_object",
                u"specification_object_id": "",
                u"missing_rep_parent_object_id": "",
                u"category": "",
                u"weight": "",
                u"priority": "",
                u"discipline": ""
            }
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"D6": spec.GetDescription(),
            u"D12": spec.Requirements.KeywordQuery(specobject_id=u'RT000000009')[0].specobject_id,
            u"D13": spec.Requirements.KeywordQuery(specobject_id=u'RT000000011')[0].specobject_id
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)

    def test_requirement_links_report(self):
        """ (R00002910) Report: Requirement Links"""
        from cs.requirements import RQMSpecification
        spec = RQMSpecification.KeywordQuery(name=u"RQM ReqIF Interface (TEST)")[0]
        req = spec.Requirements.KeywordQuery(specobject_id=u"RT000000006")[0]
        res = self._generate_report(
            {
                u"name": u"RequirementLinks",
                u"report_title": u"Anforderungsverknüpfungen",
                u"iso_code": u"de"
            },
            objects=[req],
            dlg_args={
                u"object_names": u"Spezifikation",
                u"links": u"is contained in",
                u"types": u"cdbrqm_specification",
                u"object_names2": "",
                u"links2": "",
                u"types2": "",
                u"object_names3": "",
                u"links3": "",
                u"types3": ""
            }
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"E3": u"Spezifikation",
            u"E4": u"is contained in",
            u"B12": spec.Requirements.KeywordQuery(specobject_id=u'RT000000006')[0].GetDescription(),
            u"E12": spec.GetDescription(),
            u"B13": spec.GetDescription(),
            u"E13": spec.Requirements.KeywordQuery(specobject_id=u'RT000000007')[0].GetDescription(),
            u"B14": spec.Requirements.KeywordQuery(specobject_id=u'RT000000007')[0].GetDescription(),
            u"E14": spec.Requirements.KeywordQuery(specobject_id=u'RT000000012')[0].GetDescription(),
            u"B15": spec.GetDescription(),
            u"E15": spec.Requirements.KeywordQuery(specobject_id=u'RT000000008')[0].GetDescription(),
            u"B16": spec.Requirements.KeywordQuery(specobject_id=u'RT000000008')[0].GetDescription(),
            u"E16": spec.Requirements.KeywordQuery(specobject_id=u'RT000000014')[0].GetDescription(),
            u"B17": spec.Requirements.KeywordQuery(specobject_id=u'RT000000008')[0].GetDescription(),
            u"E17": spec.Requirements.KeywordQuery(specobject_id=u'RT000000015')[0].GetDescription()
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)


    def test_requirement_rating_overview(self):
        from cs.requirements import RQMSpecification
        spec = RQMSpecification.KeywordQuery(name=u'report-test-specification')[0]
        variable_id='RQM_RATING_RQM_COMMENT_EXTERN'
        req = spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0]
        req.SetText("cdbrqm_spec_object_desc_de", """<xhtml:div>{}</xhtml:div>""".format(
            RichTextVariables.get_variable_xhtml(variable_id=variable_id)
        ))
        classification_data = classification_api.get_new_classification(
            ["RQM_RATING"], narrowed=False
        )
        cld = classification_api.get_classification(req)
        variable_value = '### test comment ###'
        classification_data['properties']['RQM_RATING_RQM_COMMENT_EXTERN'][0]['value'] = variable_value
        classification_data['properties']['RQM_RATING_RQM_RATING_VALUE'][0]['value']['de']['text_value'] = 'akzeptiert'
        classification_api.update_classification(req, classification_data)
        res = self._generate_report(
            {
                u"name": u"RequirementsRatingExport",
                u"report_title": u"Anforderungsbewertungsübersicht",
                u"iso_code": u"de"
            },
            objects=[spec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"D3": spec.name,
            u"D27": spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0].specobject_id,
            u"L27": u"### test comment ###",
            u"M27": u"akzeptiert",
            u"D28": spec.Requirements.KeywordQuery(specobject_id=u'RT000000003')[0].specobject_id,
            u"D29": spec.Requirements.KeywordQuery(specobject_id=u'RT000000005')[0].specobject_id,
            u"D30": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000002')[0].targetvalue_id,
            u"D31": spec.Requirements.KeywordQuery(specobject_id=u'RT000000001')[0].specobject_id,
            u"D32": spec.Requirements.KeywordQuery(specobject_id=u'RT000000004')[0].specobject_id,
            u"D33": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000001')[0].targetvalue_id,
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)
        self.assertIn(variable_value, tabelle1[u"C27"].value)

    def test_requirement_rating_request(self):
        from cs.requirements import RQMSpecification
        from openpyxl import load_workbook
        spec = RQMSpecification.KeywordQuery(name=u'report-test-specification')[0]
        variable_id='RQM_RATING_RQM_COMMENT_EXTERN'
        req = spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0]
        req.SetText("cdbrqm_spec_object_desc_de", """<xhtml:div>{}</xhtml:div>""".format(
            RichTextVariables.get_variable_xhtml(variable_id=variable_id)
        ))
        classification_data = classification_api.get_new_classification(
            ["RQM_RATING"], narrowed=False
        )
        cld = classification_api.get_classification(req)
        variable_value = '### test comment ###'
        classification_data['properties']['RQM_RATING_RQM_COMMENT_EXTERN'][0]['value'] = variable_value
        classification_data['properties']['RQM_RATING_RQM_RATING_VALUE'][0]['value']['de']['text_value'] = 'akzeptiert'
        classification_api.update_classification(req, classification_data)
        res = self._generate_report(
            {
                u"name": u"RequirementsRatingExport",
                u"report_title": u"Anforderungsbewertungsanfrage",
                u"iso_code": u"de"
            },
            objects=[spec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"D3": spec.name,
            u"D27": spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0].specobject_id,
            u"J27": u"### test comment ###",
            u"K27": u"akzeptiert",
            u"D28": spec.Requirements.KeywordQuery(specobject_id=u'RT000000003')[0].specobject_id,
            u"D29": spec.Requirements.KeywordQuery(specobject_id=u'RT000000005')[0].specobject_id,
            u"D30": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000002')[0].targetvalue_id,
            u"D31": spec.Requirements.KeywordQuery(specobject_id=u'RT000000001')[0].specobject_id,
            u"D32": spec.Requirements.KeywordQuery(specobject_id=u'RT000000004')[0].specobject_id,
            u"D33": spec.TargetValues.KeywordQuery(targetvalue_id=u'AT0000001')[0].targetvalue_id,
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value, val)
        self.assertIn(variable_value, tabelle1[u"C27"].value)

    def test_excel_roundtrip(self):
        from cs.requirements import RQMSpecification
        from openpyxl import load_workbook
        from cdb.objects import operations

        spec = RQMSpecification.KeywordQuery(name=u'report-test-specification')[0]
        req = spec.Requirements.KeywordQuery(specobject_id=u'RT000000000')[0]
        classification_data = classification_api.get_new_classification(
            ["RQM_RATING"], narrowed=False
        )
        classification_data['properties']['RQM_RATING_RQM_COMMENT_EXTERN'][0]['value'] = "export test"
        classification_data['properties']['RQM_RATING_RQM_RATING_VALUE'][0]['value']['de']['text_value'] = 'akzeptiert'
        classification_api.update_classification(req, classification_data)
        res = self._generate_report(
            {
                u"name": u"RequirementsRatingExport",
                u"report_title": u"Anforderungsbewertungsanfrage",
                u"iso_code": u"de"
            },
            objects=[spec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        wb = load_workbook(filename=res[0], read_only=False)
        tabelle1 = wb[wb.sheetnames[0]]
        tabelle1["L27"] = "import test"
        tabelle1["M27"] = "nicht akzeptiert"
        wb.save(res[0])
        evaluator_h = "ed981532-62b7-11ea-b0a1-34e12d2f8425"
        try:
            operations.operation(
                "cdbrqm_rating_import", spec, operations.form_input(spec, bewerter=evaluator_h, import_excel=res[0])
            )
        except ElementsError as e:
            print(e)
        cld = classification_api.get_classification(req)
        evaluators = [rating['value']['child_props']['RQM_EVALUATOR'][0]['value'] for rating in cld['properties']['RQM_RATING_RQM_RATING']]
        self.assertIn('ed981532-62b7-11ea-b0a1-34e12d2f8425', evaluators)
        for rating_entry in cld['properties']['RQM_RATING_RQM_RATING']:
            if rating_entry['value']['child_props']['RQM_EVALUATOR'][0]['value'] == 'ed981532-62b7-11ea-b0a1-34e12d2f8425':
                self.assertEqual(rating_entry['value']['child_props']['RQM_RATING_VALUE'][0]['value']['de']['text_value'], 'nicht akzeptiert')
                self.assertEqual(rating_entry['value']['child_props']['RQM_COMMENT_EXTERN'][0]['value'], 'import test')

    def test_sortorder_of_child_elements(self):
        from cs.requirements_reqif import ReqIFProfile
        from cs.requirements_reqif.reqif_import_ng import ReqIFImportNG
        from cs.requirements import RQMSpecification
        from cdb.objects import operations

        spec = operations.operation("CDB_Create", RQMSpecification, name="sortorder_test")
        profile = ReqIFProfile.ByKeys(profile_name="CIM DATABASE Standard")

        importer = ReqIFImportNG(
            specification_mappings={
                ReqIFImportNG.DEFAULT_MAPPING_KEY: spec,
                spec.reqif_id: spec
            },
            profile=profile.cdb_object_id,
            import_file=os.path.join(os.path.dirname(__file__), 'sortorder_test.reqifz'),
            logger=LOG)
        importer.imp()
        spec.Reload()
        res = self._generate_report(
            {
                u"name": u"RequirementsOverview",
                u"report_title": u"Anforderungsübersicht",
                u"iso_code": u"de"
            },
            objects=[spec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle1 = wb[wb.sheetnames[0]]
        cell_values_to_check = {
            u"B28": u"1.1",
            u"B29": u"1.2",
            u"B30": u"1.3",
            u"B31": u"1.4",
            u"B32": u"1.5",
            u"B33": u"1.6",
            u"B34": u"1.7",
            u"B35": u"1.8",
            u"B36": u"1.9",
            u"B37": u"1.10",
            u"B38": u"1.11",
            u"B39": u"1.12",
            u"B40": u"2",
            u"B41": u"2.1",
            u"B42": u"2.2",
        }
        for k, val in cell_values_to_check.items():
            self.assertEqual(tabelle1[k].value.strip(), val)

    def test_only_active_property_values_selectable(self):
        from cs.classification.catalog import Property, PropertyValue, MultilangPropertyValue
        from cs.requirements import RQMSpecification
        
        rating_value_property = Property.ByKeys(code="RQM_RATING_VALUE")
        property_values_active = [property_value.multilang_value for property_value in
                           PropertyValue.KeywordQuery(property_object_id=rating_value_property.cdb_object_id, is_active=True)]
        MultilangPropertyValue.Create(property_object_id=rating_value_property.cdb_object_id, is_active=False, multilang_value="DO NOT SHOW")
        property_values_inactive = [property_value.multilang_value for property_value in
                                  PropertyValue.KeywordQuery(property_object_id=rating_value_property.cdb_object_id,
                                                             is_active=False)]
        spec = RQMSpecification.Query()[0]
        res = self._generate_report(
            {
                u"name": u"RequirementsOverview",
                u"report_title": u"Anforderungsübersicht",
                u"iso_code": u"de"
            },
            objects=[spec],
        )
        assert isinstance(res, list) and len(res) > 0 and os.path.isfile(res[0]), u"error creating report"
        LOG.info('testing against file : %s', res[0])
        from openpyxl import load_workbook
        wb = load_workbook(filename=res[0], read_only=True)
        tabelle3 = wb[wb.sheetnames[2]]
        for row in tabelle3.rows:
            if(row[0].value == "Bewertungen"):
                continue
            self.assertIn(row[0].value, property_values_active)
            self.assertNotIn(row[0].value, property_values_inactive)

