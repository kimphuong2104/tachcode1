import copy
import logging
import zipfile
from cdb import sqlapi
from lxml import etree
from cs.classification import api
from cs.classification.tools import get_active_classification_languages
from cs.requirements.exceptions import (
    ExcelImportDataNotFound,
    ExcelImportSpecificationNotFound,
    ExcelImportTooManyTablesFound,
    ExcelImportNoMatchingTableFound
)


weight_values = {
    'de': ['nicht akzeptiert', 'Info fehlt', 'teilweise akzeptiert', 'akzeptiert', 'nicht relevant'],
    'en': ['not accepted', 'Information missing', 'partially accepted', 'accepted', 'not relevant']
}


class Rating(object):
    """Rating class to manage Excel Roundtrip and Overall rating update"""
    def __init__(self, path_to_file, rated_by_id, spec_id_context, index_context):
        from openpyxl import load_workbook
        self.par_unique_names = [
            'req_hyperlink',
            'req_comment',
            'req_rating'
        ]
        self.languages = get_active_classification_languages()
        self.path_to_file = path_to_file
        self.rated_by = rated_by_id
        self.spec_id_context = spec_id_context
        self.index_context = index_context
        self.wb = load_workbook(path_to_file)
        self.target_ws = None
        self.target_tbl = None
        self.spec_id = None
        self.spec_index = None
        self.report_lang = None
        self.req_lst = []

    @classmethod
    def get_max_revision(cls, spec_ndx):
        stmt = "max(revision) FROM cdbrqm_specification WHERE spec_id='%s'" \
               % spec_ndx
        t = sqlapi.SQLselect(stmt)

        max_revision = '0'
        if not sqlapi.SQLnull(t, 0, 0):
            max_revision = sqlapi.SQLinteger(t, 0, 0)
        return max_revision

    def trigger(self, logger=None, logger_extra_args=None):
        '''Calls all functions for Requirements Rating Excel round trip '''
        logger.info('%s: Rating update for %s', self.spec_id_context, self.rated_by, extra=logger_extra_args)
        logging.info('%s: Rating update for %s', self.spec_id_context, self.rated_by)

        # Obtain name of worksheet/table containing the requirements
        self.find_req_ws_tbl(logger, logger_extra_args)

        # Get the spec id, index and language
        self.find_spec_id_index(logger, logger_extra_args)

        # Check for empty fields in the previous procedure
        invalid_fields = []

        if self.spec_id is None or str(self.spec_id).strip() == "":
            invalid_fields.append("Specification ID")
        if self.spec_index is None or str(self.spec_index).strip() == "":
            invalid_fields.append("index")
        if self.report_lang is None or str(self.report_lang).strip() == "" or str(self.report_lang).strip() not in self.languages:
            invalid_fields.append("language")

        if len(invalid_fields) > 0:
            logger.error("The following field(s) in the Excel file are invalid: %s", invalid_fields,
                         extra=logger_extra_args)
            logger.error("The following field(s) in the Excel file are invalid: %s", invalid_fields)
            raise ExcelImportDataNotFound("cdbrqm_invalid_field")

        # Check if context specification id and the report's specification id match
        if self.spec_id_context != self.spec_id:
            logger.error("The Specification ID in the Excel file does not match the one provided by the context",
                         extra=logger_extra_args)
            logging.error("The Specification ID in the Excel file does not match the one provided by the context")
            raise ExcelImportSpecificationNotFound("cdbrqm_wrong_specification")

        # Obtain requirements
        logger.info('Retrieving requirement ratings from the Excel file', extra=logger_extra_args)
        logging.info('Retrieving requirement ratings from the Excel file')
        self.get_requirements(logger, logger_extra_args)

        # Update comment and rating on individual requirements
        self.update_req_rating_comment(logger, logger_extra_args)
        logger.info('Process executed succesfully', extra=logger_extra_args)
        logging.info('Process executed succesfully')

    @staticmethod
    def validate_rating(current_rating, languages):
        for lan in languages:
            if current_rating in weight_values[lan]:
                return True
        return False

    @staticmethod
    def get_rating_weight(rating, languages):
        for lan in languages:
            for rating_value in weight_values[lan]:
                if rating in rating_value:
                    return weight_values[lan].index(rating)
        return None

    @staticmethod
    def update_overall_rating(classification_data):
        '''Updates the overall rating of a requirement'''
        languages = get_active_classification_languages()
        # Variable to store the minumim weight
        minimum_weight = len(weight_values[languages[0]]) + 1
        if 'RQM_RATING' in classification_data["assigned_classes"]:
            # Iterate over all ratings on this requirement
            for i, _single_rating in enumerate(classification_data['properties']['RQM_RATING_RQM_RATING']):
                # Get the current rating in the first language
                if (
                    (len(classification_data['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_RATING_VALUE']) == 0) or
                    (classification_data['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_RATING_VALUE'][0]["value"] == {})
                ):
                    current_rating = None
                else:
                    current_rating = classification_data['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_RATING_VALUE'][0]["value"][languages[0]]['text_value']
                if current_rating is not None and Rating.validate_rating(current_rating, languages):
                    # get the weight of this rating. Should more languages arise, then the lists with the weights should also be modified
                    current_weight = Rating.get_rating_weight(current_rating, languages)
                    # Store if it is smaller than the smallest rating
                    if current_weight < minimum_weight:
                        minimum_weight = current_weight
            # if no error
            if minimum_weight != len(weight_values[languages[0]]) + 1:
                # Overall rating is erased after checkbox is set to true
                if len(classification_data['properties']['RQM_RATING_RQM_RATING_VALUE'][0]['value']) > 0:
                    if classification_data['properties']['RQM_RATING_RQM_RATING_VALUE'][0]['value'][languages[0]]['text_value'] != weight_values[languages[0]][minimum_weight]:
                        # Update the overall rating if its different from the previous value
                        for lan in languages:
                            classification_data['properties']['RQM_RATING_RQM_RATING_VALUE'][0]['value'][lan]['text_value'] = weight_values[lan][minimum_weight]
                else:
                    temp_data = {lan: {'text_value': weight_values[lan][minimum_weight], 'iso_language_code': lan} for lan in languages}
                    classification_data['properties']['RQM_RATING_RQM_RATING_VALUE'][0]['value'] = temp_data
            else:
                # Some ratings might have been provided by an organisation
                logging.warning('For overall rating not updated. Not all ratings were provided by a Person or are invalid')

    def update_req_rating_comment(self, logger=None, logger_extra_args=None):
        '''Updates the rating and comment on DB of a Requirement'''
        # Get the Specification object from the DB using the ID and the index
        from cs.requirements import RQMSpecification
        spec_obj = RQMSpecification.ByKeys(spec_id=self.spec_id, revision=self.spec_index, ce_baseline_id='')
        if spec_obj is None:
            logger.error('The Specification found in the Excel file does not match the latest state', extra=logger_extra_args)
            logging.error('The Specification found in the Excel file does not match the latest state')
            raise ExcelImportSpecificationNotFound("cdbrqm_specification_not_exist")
        elif( not spec_obj.CheckAccess("save")):
            logger.error('The index found in the Excel file does not match the latest state', extra=logger_extra_args)
            logging.error('The index found in the Excel file does not match the latest state')
            raise ExcelImportSpecificationNotFound("cdbrqm_wrong_index")

        #Check if the index matches the latest version ### here
        latest_index = Rating.get_max_revision(self.spec_id)
        if self.index_context != latest_index:
            logger.error("The selected specification cannot be edited anymore",
                         extra=logger_extra_args)
            logging.error("The selected specification cannot be edited anymore")
            raise ExcelImportSpecificationNotFound("cdbrqm_wrong_specification")

        # Create dictionary with key req id and values comment, rating and some day rater
        req_dict = {req['req_hyperlink']: {'req_comment': req['req_comment'], 'req_rating': req['req_rating']} for req in self.req_lst}
        # Get rating list for both languages
        rating_query = sqlapi.RecordSet2("cs_property", "code = 'RQM_RATING_VALUE'")
        ratings_set = sqlapi.RecordSet2("cs_property_value", "property_object_id = '%s'" % rating_query[0].cdb_object_id)
        # Dictionary where keys are active languages and values are a list with all the ratings
        all_ratings = {language: [getattr(rating, "multilang_value_" + language) for rating in ratings_set] for language in self.languages}
        # Iterate over the Requirements listed in that specification
        for req_obj in spec_obj.Requirements:
            # Filter out the requirements that are not in the current excel
            if req_obj['specobject_id'] in req_dict.keys():
                # Get the comment and rating and save it on the requirement object
                req_comment = req_dict[req_obj['specobject_id']]['req_comment']
                req_rating = req_dict[req_obj['specobject_id']]['req_rating']
                # get classification from requirement object
                req_class = api.get_classification(req_obj)
                # Check if the RQM_RATING class is linked, if not, link it
                if 'RQM_RATING' not in req_class["assigned_classes"]:
                    req_class = api.rebuild_classification(req_class, ['RQM_RATING'])
                # Create or update only valid ratings
                if req_rating is not None:
                    if Rating.validate_rating(req_rating, self.languages):
                        # If there is only one classification and it doesnt have a value, set comment and rating
                        if (
                            len(req_class['properties']['RQM_RATING_RQM_RATING']) == 1 and
                            req_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_RATING_VALUE'][0]['id'] is None
                        ):
                            logging.debug('The Specification id found in the Excel file does not match the latest state')
                            req_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_COMMENT_EXTERN'][0]["value"] = req_comment
                            req_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_EVALUATOR'][0]['value'] = self.rated_by
                            # Set the system language value
                            req_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_RATING_VALUE'][0]["value"][self.report_lang]['text_value'] = req_rating
                            # Set the rest of the languages if existing
                            for lan in self.languages:
                                if lan != self.report_lang:
                                    if lan in req_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_RATING_VALUE'][0]["value"]:
                                        req_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_RATING_VALUE'][0]["value"][lan]['text_value'] = all_ratings[lan][all_ratings[self.report_lang].index(req_rating)]
                        # There are more than one, or the existing one already has a classification
                        else:
                            # Iterate over existing classifications
                            found_rated_by = False
                            for i, existing_class in enumerate(req_class['properties']['RQM_RATING_RQM_RATING']):
                                # If a existing classification was rated by the same Rater, update it
                                if existing_class['value']['child_props']['RQM_EVALUATOR'][0]['value'] == self.rated_by:
                                    logging.debug('%s: Updating existing rating', req_obj['specobject_id'])
                                    req_class['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_COMMENT_EXTERN'][0]["value"] = req_comment
                                    req_class['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_EVALUATOR'][0]['value'] = self.rated_by
                                    # Set the system language value
                                    req_class['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_RATING_VALUE'][0]["value"][self.report_lang]['text_value'] = req_rating
                                    # Set the rest of the languages if existing
                                    for lan in self.languages:
                                        if lan != self.report_lang:
                                            if lan in req_class['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_RATING_VALUE'][0]["value"]:
                                                req_class['properties']['RQM_RATING_RQM_RATING'][i]['value']['child_props']['RQM_RATING_VALUE'][0]["value"][lan]['text_value'] = all_ratings[lan][all_ratings[self.report_lang].index(req_rating)]
                                    found_rated_by = True
                                    break
                            # if we iterated over the complete array without finding the same Rated_by, append a new classification
                            if found_rated_by is False:
                                # Create a new empty classification and set the rating and comment value
                                logging.debug('%s: Adding new rating', req_obj['specobject_id'])
                                empty_class = api.get_new_classification(["RQM_RATING"])
                                empty_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_COMMENT_EXTERN'][0]["value"] = req_comment
                                empty_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_EVALUATOR'][0]['value'] = self.rated_by
                                # Set the system language value
                                empty_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_RATING_VALUE'][0]["value"][self.report_lang]['text_value'] = req_rating
                                # Set the rest of the languages if existing
                                for lan in self.languages:
                                    if lan != self.report_lang:
                                        if lan in req_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_RATING_VALUE'][0]["value"]:
                                            empty_class['properties']['RQM_RATING_RQM_RATING'][0]['value']['child_props']['RQM_RATING_VALUE'][0]["value"][lan]['text_value'] = all_ratings[lan][all_ratings[self.report_lang].index(req_rating)]
                                # Append the classification to the requirements classifications
                                req_class["properties"]["RQM_RATING_RQM_RATING"].append(
                                    copy.deepcopy(empty_class["properties"]["RQM_RATING_RQM_RATING"][0])
                                )
                        # Update the Requirement in DB
                        req_class['properties']['RQM_RATING_RQM_RATING_CALCULATED'][0]['value'] = False
                        api.update_classification(req_obj, req_class)
                    else:
                        logger.warning('%s: Invalid rating found', req_obj['specobject_id'], extra=logger_extra_args)
                        logging.warning('%s: Invalid rating found', req_obj['specobject_id'])

    def get_requirements(self, logger=None, logger_extra_args=None):
        '''Updates the member list where each element is a dictionary of requirements containing the pairs of unique name and value '''
        # Load the table into an object
        data = self.target_ws[self.target_tbl.ref]
        # Boolean to skip the first row containing the column names as they are in the excel file
        first = True
        # Iterate over each row
        for row in data:
            # Ignore column names
            if first:
                first = False
            else:
                # Empty dictionary to store the pairs
                row_dict = {}
                # Iterate over each cell in the row
                for cell in row:
                    # Get the index number of the current cell
                    index = row.index(cell)
                    # Get the unique name of the column
                    key_dict = self.target_tbl.tableColumns[index].uniqueName
                    # If this key is one of the target unique names, add the pair
                    if key_dict in self.par_unique_names:
                        key_val = cell.value
                        row_dict[key_dict] = key_val
                # Add the dictionary to the list of results
                self.req_lst.append(row_dict.copy())

    def find_spec_id_index(self, logger=None, logger_extra_args=None):
        '''Stores the specification id, revision and language from the classe s a Requirements Overview excel file '''
        expected_keys = ["spec_id", "spec_index", "report_lang"]
        b_spec_id = False
        b_spec_index = False
        b_report_lang = False
        props = self.wb.custom_doc_props.props
        for prop in props:
            if 'spec_id' in prop.name:
                self.spec_id = prop.value
                expected_keys.remove("spec_id")
                b_spec_id = True
            elif 'revision' in prop.name:
                self.spec_index = prop.value
                expected_keys.remove("spec_index")
                b_spec_index = True
            elif 'cdbxml_report_lang' in prop.name:
                self.report_lang = prop.value
                expected_keys.remove("report_lang")
                b_report_lang = True
        if b_spec_id and b_spec_index and b_report_lang:
            return True

        # Fallback for older reports
        # Fixed fields
        xpath_spec_id = '/Root/SpecificationOverview/@spec_id'
        xpath_revision = '/Root/SpecificationOverview/@revision'
        xpath_report_lang = '/Root/Arguments/@cdbxml_report_lang'
        # Tags namespace
        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        # Load the workbook to a particular sheet
        ws = self.wb[self.target_ws.title]
        # File path of the
        fh = zipfile.ZipFile(self.path_to_file)
        # Get list of all tableSingleCell-n files in folder
        all_files_list = fh.namelist()
        # Filter out files that do not match the pattern
        sub_string = 'xl/tables/tableSingleCells'
        single_cell_files = [sub_elem for sub_elem in all_files_list if sub_string in sub_elem]
        # Iterate over all xml files that match the pattern
        for xml_single_cell in single_cell_files:
            # Load the Etree object with XML file
            root = etree.parse(fh.open(xml_single_cell))
            # Iterate over all parent elements with the xpath tag
            all_cells = root.xpath('/x:singleXmlCells/x:singleXmlCell', namespaces=ns)
            for parent_element in all_cells:
                location = parent_element.attrib['r']
                found_xpath = parent_element.getchildren()[0].getchildren()[0].attrib['xpath']
                if found_xpath == xpath_spec_id:
                    self.spec_id = ws[location].value
                    expected_keys.remove("spec_id")
                    b_spec_id = True
                elif found_xpath == xpath_revision:
                    self.spec_index = ws[location].value
                    expected_keys.remove("spec_index")
                    b_spec_index = True
                elif found_xpath == xpath_report_lang:
                    self.report_lang = ws[location].value
                    expected_keys.remove("report_lang")
                    b_report_lang = True
                # If all tags have been found, return values
                if b_spec_id and b_spec_index and b_report_lang:
                    return True
        # Log and raise exception
        logger.error('The Excel file does not contain the following linked field(s): %s', expected_keys, extra=logger_extra_args)
        logging.error('The Excel file does not contain the following linked field(s): %s', expected_keys)
        raise ExcelImportDataNotFound("cdbrqm_rating_no_data")

    def find_req_ws_tbl(self, logger=None, logger_extra_args=None):
        '''Matches the Worksheet and table containing the defining fields with the requirement ratings and comment'''
        logger.info('Searching for an Excel Table containing the linked fields %s', self.par_unique_names, extra=logger_extra_args)
        logging.info('Searching for an Excel Table containing the linked fields %s', self.par_unique_names)
        # List with the name of all the tables containing all of the target columns
        complete_tbl_msg_lst = []
        # Empty string for a warning message
        warning_msg = ""
        # List containing the names of the tables that have all the required column names (error)
        warning_msg_lst = []
        # Boolean flag to indicate if at least one complete table
        no_tbl_in_wb = True
        # Iterate over each worksheet in the workbook
        for ws_name in self.wb.sheetnames:
            # Get current sheet
            ws = self.wb[ws_name]
            # Iterate over all tables in the current worksheet
            for tbl in ws.tables.values():
                # Since at least one table was found in the workbook, turn off flag
                no_tbl_in_wb = False
                # Temporal list with all target table names
                par_unique_names_temp = self.par_unique_names[:]
                # Iterate over all columns in the current table
                for column in tbl.tableColumns:
                    # If the current column name is one of the target unique names, remove it
                    if column.uniqueName in par_unique_names_temp:
                        par_unique_names_temp.remove(column.uniqueName)
                # If the temporary list, all target unique names were found
                if len(par_unique_names_temp) == 0:
                    # Record the name of the table with all the target columns names
                    complete_tbl_msg_lst.append(tbl.name)
                    if len(complete_tbl_msg_lst) == 1:
                        # Save the first complete table and worksheet as members
                        self.target_ws = ws
                        self.target_tbl = tbl
                else:
                    # if at least one column name was found, send a warning and list the name of the missing columns
                    if len(par_unique_names_temp) < len(self.par_unique_names):
                        warning_msg = "Table " + tbl.name + " is missing the following column"
                        warning_msg = (warning_msg + "s: ", warning_msg + ": ")[len(par_unique_names_temp) == 1]
                        warning_msg = warning_msg + "{}".format(par_unique_names_temp)
                        warning_msg_lst.append(warning_msg)
        if len(complete_tbl_msg_lst) == 1:
            # Target woksheet and table are delivered
            logger.info("All key attributes found in table %s in Worksheet %s", self.target_tbl.name, self.target_ws.title, extra=logger_extra_args)
            logging.info("All key attributes found in table %s in Worksheet %s", self.target_tbl.name, self.target_ws.title)
            return True
        else:
            if len(complete_tbl_msg_lst) > 1:
                # More than one table containing all of the target columns was found
                logger.error('More than one table contains the column id: %s', self.par_unique_names, extra=logger_extra_args)
                logging.error('More than one table contains the column id: %s', self.par_unique_names)
                raise ExcelImportTooManyTablesFound("cdbrqm_many_matching_tables")
            else:
                if len(complete_tbl_msg_lst) == 0:
                    # No table matching the target columns was found
                    if no_tbl_in_wb:
                        warning_msg_lst.append("No tables found in the Excel file")
                    if len(warning_msg_lst) != 0:
                        for warning_item in warning_msg_lst:
                            logger.warning(warning_item, extra=logger_extra_args)
                            logging.warning(warning_item)
                    logger.error("The provided Excel file does not contain a table containing the key identifying fields", extra=logger_extra_args)
                    logging.error("The provided Excel file does not contain a table containing the key identifying fields")
                    raise ExcelImportNoMatchingTableFound("cdbrqm_no_matching_table")
            return False
