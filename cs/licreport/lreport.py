#!/usr/bin/env python
# -*- mode: python; coding: utf-8 -*-
#
# Copyright (C) 1990 - 2017 CONTACT Software GmbH
# All rights reserved.
# http://www.contact.de/
#
"""
Module cdblic.lreport

Create a license/usage report.
"""

import codecs
import datetime
import gc
import json
import logging
import os
import re
from collections import Counter, OrderedDict, defaultdict, deque
from itertools import chain
from operator import itemgetter

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError

# izip
from six.moves import xrange
from six.moves import zip as izip

from cdb import dberrors, fls, sqlapi

from . import lsession, lsite_map
from . import lstatistics as ls
from . import utils
from .generic_metrics import UserStatisticsGatherer

__docformat__ = "restructuredtext en"
__revision__ = "$Id$"

# Exported objects
__all__ = ['Collector']


class Collector(object):

    COLLECTORS = {
        'Statistics.used_modules': ls.get_used_modules,
        'Statistics.total_users': ls.get_user_count,
        'Statistics.users_by_org_id': ls.get_user_count_by_org_id,
        'Statistics.users_by_site': ls.get_user_count_by_site_id,
        'Statistics.mod_usage_by_org_id': ls.get_module_usage_by_org_id,
        'Statistics.mod_usage_by_site': ls.get_module_usage_by_site_id,
        'Statistics.mod_usage_by_user_daily':
            ls.get_module_usage_by_user_daily,
        'Statistics.start_freq_hour': ls.get_session_start_frequencies_hourly,
        'Statistics.start_freq_hour_by_org_id':
            ls.get_session_start_frequencies_hourly_by_org_id,
        'Statistics.start_freq_hour_by_site_id':
            ls.get_session_start_frequencies_hourly_by_site_id,
        'Statistics.sess_freq_by_user_daily':
            ls.get_session_frequencies_per_user_daily,
        'Statistics.start_freq_weekday':
            ls.get_session_start_frequencies_weekday_hourly,
        'Statistics.start_freq_weekday_by_site_id':
            ls.get_session_start_frequencies_weekday_hourly_by_site_id,
        'Statistics.total_hosts': ls.get_host_count,
        'Statistics.total_host_user': ls.get_host_user_count,
        'Statistics.total_host_site': ls.get_host_site_count,
        'Statistics.hosts_by_site': ls.get_host_count_by_site_id,
        'Statistics.multi_host_users': ls.get_multi_host_users,
        'Statistics.used_licenses': ls.get_license_mappings,
    }

    def __init__(self, cdb_version, customer, start_time, end_time, tables, pseudonymizer=None):
        self.start_time = start_time
        self.end_time = end_time
        self.cdb_version = cdb_version
        self.report_version = 0
        self.mapper = pseudonymizer

        timetuple = get_min_max_starttimes(tables)
        self.data = {
            'Basic.start_time': self.start_time,
            'Basic.end_time': self.end_time,
            'Basic.cdb_version': self.cdb_version,
            'Basic.customer': customer,
            'Basic.oldest_data_time': timetuple[0],
            'Basic.newest_data_time': timetuple[1],
            'Basic.report_version': self.detect_report_version(),
        }
        self.tables = tables
        self.log = logging.getLogger('cdblic.lreport.Collector')

        try:
            sqlapi.SQLselect(' COUNT(license_region_id) FROM %(cdbfls_license_site)s' % self.tables)
        except dberrors.DBError:
            self.log.debug("No license regions found.")
        else:
            self.COLLECTORS['Statistics.users_by_region'] = ls.get_user_count_by_region
            self.COLLECTORS['Statistics.mod_usage_by_region'] = ls.get_module_usage_by_region

    def detect_report_version(self):
        import glob
        import hashlib

        import pkg_resources  # part of setuptools
        version = pkg_resources.require("cs.licreport")[0].version
        h = hashlib.new('sha512')
        for f in sorted(glob.glob(os.path.join(os.path.dirname(__file__), '*.py'))):
            with open(f, 'rb') as fd:
                h.update(fd.read())
        hversion = h.hexdigest()
        return "%s [%s]" % (version, hversion[:20])

    def collect(self):
        self.log.info("Collecting info for %s to %s",
                      self.start_time, self.end_time)
        for key, func in self.COLLECTORS.items():
            self.log.debug("Start collecting %s...", key)
            self.data[key] = func(self.start_time, self.end_time, self.tables,
                                  pseudonymizer=self.mapper)
            self.log.info("Collected %s", key)

        self.log.info("Collecting generic metrics about registered users")
        ug = UserStatisticsGatherer(self.cdb_version)
        ug.collect(self.data)

        # add the used org_id -> org name mappings
        # add the used site_id -> site name mappings
        # add the used mno -> license name mappings

    def dump(self, fileobj):
        self.log.debug("Dumping %d datasets to %s",
                       len(self.data),
                       repr(fileobj))
        json.dump(self.data, codecs.getwriter('utf-8')(fileobj), ensure_ascii=False)
        self.log.debug("Finished dump")


MSGS_DE = {
    u"Users per Site": u"Nutzer und Standorte",
    u"Customer": u"Kunde",
    u"Software Version": u"Software Version",
    u"Report Time Interval": u"Report Zeitraum",
    u"Number of unique User in the measurement interval":
        u"Anzahl Nutzer im Messzeitraum",
    u"Number of registered Users": u"Anzahl Nutzer im System",
    u"Dynamic User Registration (LDAP) activ?": u"LDAP aktiv",
    u"yes": u"Ja",
    u"no": u"Nein",
    u"User by Sites": u"Nutzer pro Standort",
    u"Zones": u"Zonen",
    u"Users by Zones": u"Nutzer pro Zone",
    u"Users Total": u"Nutzer Total",
    u"Users %": u"Nutzer %",
    u'License-Name': u"Lizenzname",
    u'User Distribution': u"Nutzerverteilung",
    u'User Distribution by Zone': u"Nutzerverteilung auf Zonen",
    u'Overview': u"Nutzer_und_Standorte",
    u'Usage Daily': u"Tagesnutzung",
    u'Number of Users by daily module usage': u"Lizenznutzung pro Tag",
    u'Number of Starts per Hour (averaged)':
        u"Sitzungsstarts pro Stunde (gemittelt)",
    u"Concurrent Max. Sessions":
        u"Maximale Nutzung pro Standort",
    u"Concurrent Max. Sessions by Zones":
        u"Maximale Nutzung pro Zone",
    u"License": u"Lizenz",
    u"Sum": u"Summe",
    u"Average": u"Mittelwert",
    u"Conc. Sess. Max by Site": u"Lizenzbedarf (Summe Standorte)",
    u"Start Time Daily": u"Startzeiten",
    u"User Site Mapping": u"Benutzer und Standorte",
    u"User ID": u"Personalnummer",
    u"Country": u"Land",
    u"Site": u"Standort",
    u"User Sites": u"User_Standorte",
    u"Number of unique users of this license module on the given day.":
    u"Anzahl der verschiedenen Nutzer, welche eine bestimmte Lizenz am "
    u"angegeben Tag benutzt haben.",
    u"Sessions": u"Sitzungen",
    u'Max Conc.': u"Parallel",
    u'% Conc.': u"% Parallel",
    u'Site Conc.': u"Parallel",
    u'License Type': u'Lizenztyp',
    u'Description': u'Beschreibung',
    u'Time of Day': u'Tageszeit',
    u'Starts per Hour': u'Starts pro Stunde',
    u'Start Times': u'Startzeiten',
    u'Working Hours': u'Werktag',
    u'Session Distribution': u'Session-Verteilung',
    u'Session Duration in Hours': u'Sitzungsdauer in Stunden',
    u'Session Count': u'Anzahl Sitzungen',
    u'Session Percent': u'Prozent Sitzungen',
    u'License Number': u'Lizenznummer',
    u'License Name': u'Lizenzname',
    u'License Names': u'Lizenznamen',
    u'Failed License Allocations': u'Erfolglose Versuche eine Lizenz zu nutzen',
    u'Inventory': u'Bestand',
    u'Quota': u'Quote',
}


def _(msg):
    """Translation helper for i18n"""
    return MSGS_DE.get(msg, msg)


class XLSXFormatter(object):

    def __init__(self, site_mapper, version):
        self.log = logging.getLogger('cdblic.lreport.XLSXFormatter')
        self.wb = Workbook()
        self.mapper = site_mapper
        self.version = version if version else 10
        self.setup_styles()
        self.patch_targets = {}
        self.ordered_concurrent_lics = None
        self.start = None
        self.end = None

    @staticmethod
    def mno_to_name(mno):
        if not mno:
            return mno
        mno = mno.replace('-WIN', '')
        return fls.get_license_label(mno)

    def site_to_name(self, site_id):
        return self.mapper.site2name(site_id)

    def setup_styles(self):
        # fonts
        font_red = Font(color=Color(indexed=2))
        font_bold = Font(bold=True)
        font_heading = Font(size=18, bold=True)

        # alignments
        align_rotate_90 = Alignment(text_rotation=90)
        align_center_hv = Alignment(horizontal="center", vertical="center")

        # borders
        border_bottom = Border(bottom=Side(border_style='medium'))
        border_right = Border(right=Side(border_style='medium'))
        border_bottom_right = Border(bottom=Side(border_style='medium'),
                                     right=Side(border_style='medium'))

        # add styles
        styles = [
            NamedStyle(name=u'DATE', number_format="YYYY.MM.DD"),
            NamedStyle(name=u'TIMESTAMP', number_format="YYYY.MM.DD hh:mm:ss"),
            NamedStyle(name=u'TIMESTAMP_DE',
                       number_format="DD.MM.YYYY hh:mm:ss"),
            NamedStyle(name=u'DURATION', number_format="[hh]:mm:ss"),
            NamedStyle(name=u'TIME', number_format="hh:mm"),
            NamedStyle(name=u'HOURS', number_format="hh:mm"),
            NamedStyle(name=u'ONEFPDIGIT', number_format="0.0"),
            NamedStyle(name=u'RED_WARN', font=font_red),
            NamedStyle(name=u'PERCENT', number_format="0.0%"),
            NamedStyle(name=u'HEADING', font=font_bold),
            NamedStyle(name=u'HEADING_BORDERB', font=font_bold,
                       border=border_bottom),
            NamedStyle(name=u'HEADING_BORDERR', font=font_bold,
                       border=border_right),
            NamedStyle(name=u'ROTATED_HEADING', alignment=align_rotate_90),
            NamedStyle(name=u'PAGEHEADING', font=font_heading,
                       alignment=align_center_hv),
            NamedStyle(name=u'BORDERB', border=border_bottom),
            NamedStyle(name=u'BORDERR', border=border_right),
            NamedStyle(name=u'BORDERBR', border=border_bottom_right),
            NamedStyle(name=u'CENTERED', alignment=align_center_hv),
            NamedStyle(name=u'CENTEREDPCT',
                       alignment=align_center_hv, number_format="0.00%"),
            NamedStyle(name=u'BORDERR_CENTERED', border=border_right,
                       alignment=align_center_hv),
        ]
        for s in styles:
            self.wb.add_named_style(s)

    @staticmethod
    def mask_value(value):
        # remove unicode control stuff invalid in xml
        bad_xml = re.compile(
            u'[\x00-\x08\x0b\x0c\x0e-\x1F\uD800-\uDFFF\uFFFE\uFFFF]')
        return bad_xml.sub('?', value)

    def set_cell(self, ws, row, column, value, style=None):
        c = ws.cell(row=row, column=column)
        try:
            c.value = value
        except IllegalCharacterError:
            c.value = self.mask_value(value)
        if style:
            c.style = style
        return c

    def format_license_info(self, ws, dataset):
        # License data
        licenses = dataset['Statistics.used_licenses']

        # Sheet heading
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=3)
        self.set_cell(ws, 1, 1, _(u'License Names'), 'PAGEHEADING')

        # Format basic usage metrics
        crow = 3
        # Header row
        headings = [u'License Number', u'License Name']
        for col, colname in enumerate(headings, 1):
            self.set_cell(ws, crow, col, colname, 'HEADING')
        crow += 1

        # List all licenses
        for mno in licenses.keys():
            self.set_cell(ws, crow, 1, mno)
            self.set_cell(ws, crow, 2, self.mno_to_name(mno))
            crow += 1

        # List licenses contained in packages
        crow += 2
        ccol = 1
        for mno, sublics in licenses.items():
            if sublics:
                row_bck = crow
                self.set_cell(ws, row_bck, ccol, self.mno_to_name(mno), u'HEADING_BORDERB')
                for lic in sublics:
                    row_bck += 1
                    self.set_cell(ws, row_bck, ccol, self.mno_to_name(lic))
                ws.column_dimensions[get_column_letter(ccol)].width = 25
                ccol += 1

        return crow

    def layout_users_by_region(self, ws, dataset, startrow,
                               total_user_row, total_user_column):
        crow = startrow
        # Users-by-Region
        self.set_cell(ws, crow, 10, _(u"Users by Zones"), 'HEADING')
        crow += 2

        self.set_cell(ws, crow, 10, _(u"Zones"), 'HEADING')
        self.set_cell(ws, crow, 11, _(u"Users Total"), 'HEADING')
        self.set_cell(ws, crow, 12, _(u"Users %"), 'HEADING')
        crow += 1

        data_rows_start_region = crow
        dsr = dataset['Statistics.users_by_region']
        for region, count in sorted(dsr.items(),
                                    key=itemgetter(1), reverse=True):
            if not region:
                region = "Unknown"
            self.set_cell(ws, crow, 10,
                          region, 'HEADING')
            self.set_cell(ws, crow, 11, int(count))
            self.set_cell(ws, crow, 12, "=%s%s/$%s$%s" % (
                get_column_letter(11), crow, get_column_letter(total_user_column),
                total_user_row), 'PERCENT')
            crow += 1
        data_rows_end_region = crow - 1
        dim = ws.column_dimensions['J']
        dim.width = 30.0
        self.layout_user_by_region_chart(ws, data_rows_start_region,
                                         data_rows_end_region)

    def format_header(self, ws, dataset):
        self.log.debug("Creating header")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
        self.set_cell(ws, 1, 1, _(u"Users per Site"), 'PAGEHEADING')

        crow = 3
        self.set_cell(ws, crow, 1, _(u"Customer"), 'HEADING')
        self.set_cell(ws, crow, 2, dataset['Basic.customer'])
        self.set_cell(ws, crow, 3, _(u"Report Time Interval"), 'HEADING')
        self.set_cell(ws, crow, 4, _(u"Start"), 'HEADING')
        self.set_cell(ws, crow, 5, dataset['Basic.start_time'])
        self.set_cell(ws, crow, 7, _(u"Available Dataset"), 'HEADING')
        self.set_cell(ws, crow, 8, _(u"Oldest data"), 'HEADING')
        self.set_cell(ws, crow, 9, dataset['Basic.oldest_data_time'])

        crow += 1
        self.set_cell(ws, crow, 1, _(u"Software Version"), 'HEADING')
        self.set_cell(ws, crow, 2, dataset['Basic.cdb_version'])
        self.set_cell(ws, crow, 4, _(u"End"), 'HEADING')
        self.set_cell(ws, crow, 5, dataset['Basic.end_time'])
        self.set_cell(ws, crow, 8, _(u"Newest Data"), 'HEADING')
        self.set_cell(ws, crow, 9, dataset['Basic.newest_data_time'])

        crow += 1
        self.set_cell(ws, crow, 1, _(u"Report Version"), 'HEADING')
        self.set_cell(ws, crow, 2, dataset['Basic.report_version'])

        crow += 2

        self.set_cell(ws, crow, 1,
                      _(u"Number of unique User in the measurement interval"),
                      'HEADING')
        self.set_cell(ws, crow, 2,
                      int(dataset['Statistics.total_users']['user_count']))

        total_user_column = 2
        total_user_row = crow

        crow += 1
        self.set_cell(ws, crow, 1, _(u"Number of registered Users"), 'HEADING')
        self.set_cell(ws, crow, 2, int(dataset['UserMetrics.active_users']))

        crow += 1

        self.set_cell(ws, crow, 1,
                      _(u"Dynamic User Registration (LDAP) activ?"),
                      'HEADING')
        self.set_cell(ws, crow, 2,
                      _(dataset['UserMetrics.has_dynamic_creation']))

        crow += 3
        self.set_cell(ws, crow, 1, _(u"User by Sites"), 'HEADING')

        crow += 2
        self.set_cell(ws, crow, 1, _(u"Sites"), 'HEADING')
        self.set_cell(ws, crow, 2, _(u"Users Total"), 'HEADING')
        self.set_cell(ws, crow, 3, _(u"Users %"), 'HEADING')

        crow += 1
        # Users-by-Site
        data_rows_start = crow
        dss = dataset['Statistics.users_by_site']
        for site_id, count in sorted(dss.items(),
                                     key=itemgetter(1), reverse=True):
            self.set_cell(ws, crow, 1,
                          self.site_to_name(site_id), 'HEADING')
            self.set_cell(ws, crow, 2, int(count))
            self.set_cell(ws, crow, 3, "=%s%s/$%s$%s" % (
                'B', crow, get_column_letter(total_user_column),
                total_user_row), 'PERCENT')
            crow += 1
        data_rows_end = crow - 1
        dim = ws.column_dimensions['A']
        dim.width = 30.0

        # Analysis by region
        if self.mapper.have_regions():
            self.layout_users_by_region(ws, dataset, data_rows_start - 3,
                                        total_user_row, total_user_column)

        self.layout_user_by_site_chart(ws, data_rows_start, data_rows_end)

        padding = data_rows_end - data_rows_start
        if padding < 16:
            crow += (16 - padding)

        # return max header row
        return crow

    @staticmethod
    def layout_user_by_site_chart(ws, data_rows_start, data_rows_end):
        chart = PieChart()
        chart.width = 10
        chart.height = 10
        labels = Reference(ws, min_col=1, min_row=data_rows_start, max_row=data_rows_end)
        data = Reference(ws, min_col=2, min_row=data_rows_start - 1, max_row=data_rows_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = _(u"User Distribution")
        ws.add_chart(chart, "D13")

    @staticmethod
    def layout_user_by_region_chart(ws, data_rows_start, data_rows_end):
        chart = PieChart()
        chart.width = 10
        chart.height = 10
        labels = Reference(ws, min_col=10, min_row=data_rows_start, max_row=data_rows_end)
        data = Reference(ws, min_col=11, min_row=data_rows_start - 1, max_row=data_rows_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = _(u"User Distribution by Zone")
        ws.add_chart(chart, "M13")

    def format_basic_usage_metrics(self, ws, startrow, dataset):
        self.log.debug("Creating basic usage metric")
        # Format basic usage metrics
        crow = startrow
        # header rows
        headings = [
            _(u'License-Name'),
            _(u'Total'),
            _(u'Max Conc.'),
            _(u'% Conc.'),
            _(u'Inventory'),
            _(u'Quota'),
        ]

        ds = dataset['Statistics.mod_usage_by_site']

        s = set(chain.from_iterable(ds.values()))
        all_sites = sorted([(self.site_to_name(sid), sid) for sid in s],
                           key=itemgetter(0))

        pre_headings_len = len(headings)

        # Make Room for Back References to Module Usage
        # per Site/Total here later
        for s, id_ in all_sites:
            headings.extend([_(u'Total'), _(u'Site Conc.'), _(u'% Conc.')])

        # Add Site Names on top
        for col, (s, id_) in enumerate(all_sites, 0):
            colref = col * 3 + pre_headings_len + 1
            self.set_cell(ws, crow, colref, s, 'HEADING')
        crow += 1

        # Add other Headings below
        for col, colname in enumerate(headings, 1):
            self.set_cell(ws, crow, col, colname, 'HEADING')
        crow += 1

        mno_order = sorted([(mno, sum(data.values()))
                            for mno, data in ds.items()],
                           key=itemgetter(1), reverse=True)

        self.patch_targets = {}
        for mno, total in mno_order:
            data = ds[mno]
            self.patch_targets[mno] = {}
            self.set_cell(ws, crow, 1, self.mno_to_name(mno))
            self.set_cell(ws, crow, 2, total)
            self.patch_targets[mno]['__max__'] = (3, crow)
            self.set_cell(ws, crow, 4,
                          "=%s%d/%s%d" % ('C', crow, 'B', crow), 'PERCENT')
            # License inventory, not known usually
            self.set_cell(ws, crow, 5, 0)
            self.set_cell(ws, crow, 6,
                          "=%s%d/%s%d" % ('E', crow, 'C', crow), 'PERCENT')

            col = 7
            for site, site_id in all_sites:
                val = data.get(site_id)
                if val:
                    self.set_cell(ws, crow, col, int(data[site_id]))
                    self.patch_targets[mno][site] = (col + 1, crow)
                    self.set_cell(ws, crow, col + 2,
                                  "=%s%d/%s%d" % (get_column_letter(col + 1),
                                                  crow,
                                                  get_column_letter(col),
                                                  crow),
                                  'PERCENT')
                col += 3
            crow += 1
        return crow

    def patch_max_per_site(self, ws, lic_max_references, sheet_name):
        """Patch the calculated max values into the first page"""

        for mno, sitedata in self.patch_targets.items():
            for site, colref in sitedata.items():
                try:
                    target = lic_max_references[mno][site]
                except KeyError:
                    key = 0
                else:
                    key = "='%s'!%s" % (sheet_name, target)
                self.set_cell(ws, colref[1], colref[0], key)

    def format_hosts(self, ws, dataset):
        self.log.debug("Creating hosts dataset.")
        crow = 1
        ws.merge_cells(start_row=crow, start_column=1,
                       end_row=crow + 1, end_column=5)

        self.set_cell(ws, crow, 1, _(u"Hostname Analysis"), 'PAGEHEADING')
        crow += 2

        self.set_cell(ws, crow, 1, _(u'Unique Hostnames'), 'HEADING')
        self.set_cell(ws, crow, 2,
                      int(dataset['Statistics.total_hosts']['host_count']))
        crow += 1

        self.set_cell(ws, crow, 1, _(u'Unique Users'), 'HEADING')
        self.set_cell(ws, crow, 2,
                      int(dataset['Statistics.total_users']['user_count']))
        crow += 1

        self.set_cell(ws, crow, 1, _(u'Unique Hostname/User Pairs'), 'HEADING')
        val = int(dataset['Statistics.total_host_user']['host_user_count'])
        self.set_cell(ws, crow, 2, val)
        crow += 1

        self.set_cell(ws, crow, 1, _(u'Unique Hostname/Site Pairs'), 'HEADING')
        self.set_cell(ws, crow, 2,
                      int(dataset['Statistics.total_host_site']['host_site']))
        crow += 2

        self.set_cell(ws, crow, 1, _(u'Unique Hostnames per Site'), 'HEADING')

        crow += 1
        self.set_cell(ws, crow, 1, _(u'Site'), 'HEADING')
        self.set_cell(ws, crow, 2, _(u'Hosts'), 'HEADING')
        crow += 1

        first_row = crow
        ds = dataset['Statistics.hosts_by_site']
        for site_id, count in sorted(ds.items(),
                                     key=itemgetter(1), reverse=True):
            self.set_cell(ws, crow, 1, self.site_to_name(site_id))
            self.set_cell(ws, crow, 2, count)
            crow += 1
        last_row = crow - 1

        first = '%s%s' % (get_column_letter(2), first_row)
        last = '%s%s' % (get_column_letter(2), last_row)
        self.set_cell(ws, crow, 2, "=SUM(%s:%s)" % (first, last))

        crow += 1
        self.set_cell(ws, crow, 1,
                      _(u'Users with more than one host'), 'HEADING')
        crow += 1
        self.set_cell(ws, crow, 1, _(u'User'), 'HEADING')
        self.set_cell(ws, crow, 2, _(u'No. of Hosts'), 'HEADING')
        crow += 1
        host_items = list(dataset['Statistics.multi_host_users'].items())
        host_items = sorted(host_items, key=itemgetter(1), reverse=True)
        for uname, count in host_items:
            # pseudonymize
            self.set_cell(ws, crow, 1, self.mapper.map_uname(uname))
            self.set_cell(ws, crow, 2, count)
            crow += 1

        dim = ws.column_dimensions['A']
        dim.width = 30.0

        return crow

    @staticmethod
    def _get_mno_order_daily_usage(dataset):
        ds = dataset['Statistics.mod_usage_by_user_daily']
        # sort by average daily usage, so find average usage by mno and day
        usage = defaultdict(int)
        for row in ds.values():
            for mno, count in row.items():
                usage[mno] += count
        avg_usage = {}
        for mno in usage.keys():
            avg_usage[mno] = usage[mno] / (1.0 * len(ds))
        avg_mnos = sorted(avg_usage.items(), key=itemgetter(1), reverse=True)

        mno_header_list = [mno for (mno, count) in avg_mnos]
        return mno_header_list

    def format_mod_usage_daily(self, ws, dataset):
        self.log.debug("Creating daily mod usage")

        # prepare data summaries
        mno_header_list = self._get_mno_order_daily_usage(dataset)

        # layout
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=2, end_column=8)
        self.set_cell(ws, 1, 1, _(u'Number of Users by daily module usage'),
                      'PAGEHEADING')

        ws.merge_cells(start_row=3, start_column=1,
                       end_row=4, end_column=30)
        self.set_cell(ws, 3, 1,
                      _(u"Number of unique users of this license "
                        u"module on the given day."))
        crow = 8

        # write the header
        for col, mno in enumerate(mno_header_list, 2):
            self.set_cell(ws, crow, col, self.mno_to_name(mno),
                          'ROTATED_HEADING')
        crow += 1

        start_row = crow
        # write the daily list
        ds = dataset['Statistics.mod_usage_by_user_daily']
        for date_key in sorted(ds.keys()):
            self.set_cell(ws, crow, 1, date_key, 'DATE')
            datarow = ds[date_key]
            for col, mno in enumerate(mno_header_list, 2):
                user_count = int(datarow.get(mno, 0))
                self.set_cell(ws, crow, col, user_count)
            crow += 1
        end_row = crow - 1

        # write the averages formula
        self.set_cell(ws, crow, 1, _(u'Average'), 'HEADING')
        for col, mno in enumerate(mno_header_list, 2):
            self.set_cell(ws, crow, col, "=AVERAGE(%s%s:%s%s)" % (
                get_column_letter(col), start_row,
                get_column_letter(col), end_row))

        for col, mno in enumerate(mno_header_list, 2):
            ws.column_dimensions[get_column_letter(col)].width = 5.0

        return crow

    def format_start_time_daily(self, ws, dataset):
        self.log.debug("Creating start time table")

        data_site = dataset['Statistics.start_freq_weekday_by_site_id']
        data_total = dataset['Statistics.start_freq_weekday']

        # get all occuring site IDs
        all_site_ids = set(chain.from_iterable(s.keys()
                                               for s in data_site.values()))

        all_sites = sorted([(self.site_to_name(sid), sid)
                            for sid in all_site_ids],
                           key=itemgetter(0))

        crow = 1
        ws.merge_cells(start_row=crow, start_column=1,
                       end_row=crow + 1, end_column=6)
        self.set_cell(ws, crow, 1,
                      _(u'Number of Starts per Hour (averaged)'),
                      'PAGEHEADING')
        crow += 25
        self.set_cell(ws, crow, 2, _(u'Total'), 'HEADING')
        for col, site in enumerate(all_sites, 3):
            self.set_cell(ws, crow, col, site[0], 'HEADING')
        # crow += 1

        self.set_cell(ws, crow, 1, _(u'Working Hours'), 'HEADING')
        crow += 1

        for hours in xrange(0, 24):
            self.set_cell(ws, crow, 1, "%02d:00" % hours, 'HOURS')
            key = "%02d" % hours
            self.set_cell(ws, crow, 2, data_total.get(key, 0), 'ONEFPDIGIT')
            site_data = data_site.get(key, {})
            for col, site in enumerate(all_sites, 3):
                self.set_cell(ws, crow, col, site_data.get(site[1], 0),
                              'ONEFPDIGIT')
            crow += 1

        self.layout_start_time_chart(ws, crow - 24, crow - 1, all_sites)
        return crow

    @staticmethod
    def layout_start_time_chart(ws, start_row, end_row, all_sites):
        chart = LineChart()
        chart.title = _(u"Start Times")
        chart.y_axis.title = _(u'Starts per Hour')
        chart.x_axis.title = _(u'Time of Day')
        chart.width = 20
        chart.height = 11
        chart.style = 10

        max_col = 2 + len(all_sites)
        data = Reference(ws, min_col=2, max_col=max_col, min_row=start_row - 1, max_row=end_row)
        labels = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        for series in chart.series:
            series.smooth = True

        ws.add_chart(chart, 'A4')

    def format_license_sets_user(self, ws, sessions):
        self.log.debug("Creating license set user table")
        user_licsets = lsession.count_license_sets_by_user(sessions)
        user_licset_items = sorted(((k, v) for k, v in user_licsets.items()),
                                   key=itemgetter(1), reverse=True)
        crow = 1
        self.set_cell(ws, crow, 1, _(u'License Sets By Users'),
                      'HEADING')
        crow += 1
        self.set_cell(ws, crow, 1, _(u'User Count'), 'HEADING')
        self.set_cell(ws, crow, 2, _(u'Licenses'), 'HEADING')
        crow += 1

        for k, v in user_licset_items:
            self.set_cell(ws, crow, 1, v)
            self.set_cell(ws, crow, 2, u", ".join(k))
            crow += 1

        crow += 1
        self.log.debug("Finished license set user table")
        return crow

    def format_license_sets_site(self, ws, sessions):
        self.log.debug("Creating license set site table")
        site_licsets = lsession.count_license_sets_by_site_and_user(
            sessions, lambda u: self.mapper.map_user(u)[1])

        # keys are [mno][site] = count

        # find all sites
        all_sites = set()
        for siteval in site_licsets.values():
            for k in siteval.keys():
                all_sites.add(k)

        all_sites = list(sorted(all_sites))

        # find all sets
        all_mno_sets = sorted(site_licsets.keys())

        # header
        crow = 1
        self.set_cell(ws, crow, 1, _(u'License Sets By Site'),
                      'HEADING')
        crow += 1
        self.set_cell(ws, crow, 1, _(u'Licenses'), 'HEADING')

        for col, site in enumerate(all_sites, 2):
            self.set_cell(ws, crow, col, site, 'HEADING')
        crow += 1

        # data
        for lset in all_mno_sets:
            self.set_cell(ws, crow, 1, u", ".join(lset), 'HEADING')
            for col, site in enumerate(all_sites, 2):
                self.set_cell(ws, crow, col, site_licsets[lset].get(site, 0))
            crow += 1

        crow += 1
        self.log.debug("Finished license set site table")
        return crow

    def _daily_intervals(self):
        # add daily interval points
        dt_points = [self.start]
        next_day = datetime.datetime.fromordinal(self.start.toordinal() + 1)
        if next_day > self.end:
            dt_points.append(self.end)
        else:
            dt_points.append(next_day)
            n = 1
            while 1:
                n += 1
                next_day = datetime.datetime.fromordinal(
                    self.start.toordinal() + n)
                if next_day > self.end:
                    break
                dt_points.append(next_day)
            dt_points.append(self.end)

        return [d.strftime("%Y.%m.%d %H:%M:%S") for d in dt_points]

    def _weekly_intervals(self, firstisoweekday=1):
        # add weekly (default monday=1 starts) interval points
        dt_points = [self.start]
        next_day = datetime.datetime.fromordinal(self.start.toordinal() + 1)
        if next_day > self.end:
            dt_points.append(self.end)
        else:
            n = 0
            while 1:
                n += 1
                next_day = datetime.datetime.fromordinal(
                    self.start.toordinal() + n)
                if next_day > self.end:
                    break
                # is it a monday?
                if next_day.isoweekday() == firstisoweekday:
                    dt_points.append(next_day)

            dt_points.append(self.end)

        return [d.strftime("%Y.%m.%d %H:%M:%S") for d in dt_points]

    def format_daily_concurrent_sessions(self, ws, sessions):
        self.log.debug("Creating daily concurrent sessions table")

        event_queue = lsession.build_event_queue(
            sessions,
            lambda uname: self.mapper.map_user(uname)[1])
        points = self._daily_intervals()
        results = lsession.count_lic_events_grouped(event_queue, points)

        # get all sites/groups
        group_set = set()
        for gcount, gmaxima, gtime_maxima in results.values():
            for group in gmaxima.keys():
                group_set.add(group)

        group_headers = sorted(group_set)

        # layout the dataset
        crow = 1
        self.set_cell(ws, crow, 1, _(u'Concurrent Daily Sessions'),
                      'HEADING')
        crow += 1
        for time_key, counts in results.items():
            gcount, gmaxima, gtime_maxima = counts
            start_time, end_time = time_key

            istart = utils.parse_timestamp(start_time)

            self.set_cell(ws, crow, 1, istart)
            crow += 1

            self.set_cell(ws, crow + 1, 2, _(u"License"), 'HEADING')

            # layout days from/licenses from top to bottom,
            # sites from left to right
            col = 3
            for group in group_headers:
                self.set_cell(ws, crow, col, group, 'HEADING')
                self.set_cell(ws, crow + 1, col, _(u'Sessions'), 'HEADING')
                self.set_cell(ws, crow + 1, col + 1, _(u'Maxtime'), 'HEADING')
                col += 2

            by_group = chain.from_iterable(gmaxima.values())
            all_lics = sorted(set((s for s in by_group)))

            crow += 2
            col = 2
            for rownum, lic in enumerate(all_lics, crow):
                self.set_cell(ws, rownum, col, self.mno_to_name(lic))

            col = 3
            for group in group_headers:
                maxima = gmaxima.get(group, None)
                if maxima is not None:
                    time_maxima = gtime_maxima[group]

                    for rownum, lic in enumerate(all_lics, crow):
                        mval = maxima.get(lic, 0)
                        mtime = None
                        if mval:
                            self.set_cell(ws, rownum, col, mval)
                            mtime = time_maxima.get(lic)
                        # only add time for lics > 0
                        if mtime and mval:
                            self.set_cell(ws, rownum, col + 1,
                                          utils.parse_timestamp(mtime),
                                          'TIME')
                col += 2
            crow += len(all_lics)

    def format_daily_concurrent_sessions_max(self, ws, sessions, lic_references):
        self.log.debug("Creating max concurrent sessions table")

        event_queue = lsession.build_event_queue(
            sessions,
            lambda uname: self.mapper.map_user(uname)[1])
        points = [d.strftime("%Y.%m.%d %H:%M:%S") for d in [self.start,
                                                            self.end]]
        results = lsession.count_lic_events_grouped(event_queue, points)

        # get all sites/groups
        group_set = set()
        for gcount, gmaxima, gtime_maxima in results.values():
            for group in gmaxima.keys():
                group_set.add(group)

        group_headers = sorted(group_set)

        crow = 1
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
        self.set_cell(ws, 1, 1, _(u"Concurrent Max. Sessions"), 'PAGEHEADING')
        ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=10)
        crow += 5

        for time_key, counts in results.items():
            gcount, gmaxima, gtime_maxima = counts
            start_time, end_time = time_key
            break

        self.set_cell(ws, crow + 1, 1, _(u"License"), 'HEADING')
        self.set_cell(ws, crow + 1, 2, _(u"Sum"), 'HEADING')

        col = 3
        for group in group_headers:
            self.set_cell(ws, crow, col, group, 'HEADING')
            self.set_cell(ws, crow + 1, col, _(u'Maximum'), 'HEADING')
            self.set_cell(ws, crow + 1, col + 1, _(u'Maxtime'), 'HEADING')
            col += 2

        by_group = chain.from_iterable(gmaxima.values())
        all_lics = sorted(set((s for s in by_group)))

        # calculate summed maxima to determine order of lics...
        licsum = []
        for lic in all_lics:
            values = []
            for group in group_headers:
                maxima = gmaxima.get(group, {})
                values.append(maxima.get(lic, 0))
            licsum.append((lic, sum(values)))

        ordered_lics = [k for k, v in sorted(licsum, key=itemgetter(1),
                                             reverse=True)]
        # Store order so we can reuse it in the zone based list
        self.ordered_concurrent_lics = ordered_lics

        crow += 2
        col = 1
        for rownum, lic in enumerate(ordered_lics, crow):
            self.set_cell(ws, rownum, col, self.mno_to_name(lic))

        col = 3

        timestamp_cols = []
        for group in group_headers:
            timestamp_cols.append(col + 1)
            maxima = gmaxima.get(group, None)
            if maxima is not None:
                time_maxima = gtime_maxima[group]

                for rownum, lic in enumerate(ordered_lics, crow):
                    sum_fields = [
                        "%s%s" % (get_column_letter(i), rownum)
                        for i in xrange(3, 3 + len(group_headers) * 2, 2)
                    ]
                    self.set_cell(ws, rownum, 2,
                                  "=SUM(%s)" % ",".join(sum_fields))

                    # Store references to maxima, for use in first page
                    lic_references[lic] = {'__max__': "%s%d" % (get_column_letter(2), rownum)}
                    for group, ref in zip(group_headers, sum_fields):
                        lic_references[lic][group] = ref

                    mval = maxima.get(lic, 0)
                    if mval:
                        self.set_cell(ws, rownum, col, mval)
                        mtime = time_maxima.get(lic)
                        # only add time for lics > 0
                        if mtime and mval:
                            self.set_cell(ws, rownum, col + 1,
                                          utils.parse_timestamp(mtime),
                                          'TIMESTAMP_DE')
            col += 2
        crow += len(all_lics)
        dim = ws.column_dimensions['A']
        dim.width = 30.0
        for colid in timestamp_cols:
            dim = ws.column_dimensions[get_column_letter(colid)]
            dim.width = 20.0

    def format_daily_concurrent_sessions_reg_max(self, ws, sessions, lic_references, sheet_name):
        self.log.debug("Creating zone based max concurrent session table")

        # All regions
        regions = set(self.mapper.site2region.values())
        regions.add('Unknown')
        region_headers = sorted(regions)

        crow = 1
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
        self.set_cell(ws, 1, 1, _(u"Concurrent Max. Sessions by Zones"), 'PAGEHEADING')
        ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=10)
        crow += 5

        self.set_cell(ws, crow + 1, 1, _(u"License"), 'HEADING')
        self.set_cell(ws, crow + 1, 2, _(u"Sum"), 'HEADING')

        col = 3
        for region in region_headers:
            self.set_cell(ws, crow, col, region, 'HEADING')
            self.set_cell(ws, crow + 1, col, _(u'Maximum'), 'HEADING')
            col += 1

        # The order is identical to the site based stats for easier comparision
        # It must be, due to the sum over all.
        crow += 2
        col = 1
        for rownum, lic in enumerate(self.ordered_concurrent_lics, crow):
            self.set_cell(ws, rownum, col, self.mno_to_name(lic))

        # Totals over all regions
        col = 2
        for rownum, lic in enumerate(self.ordered_concurrent_lics, crow):
            sum_fields = [
                "%s%s" % (get_column_letter(i), rownum)
                for i in xrange(3, len(region_headers) + 3)
            ]
            self.set_cell(ws, rownum, 2,
                          "=SUM(%s)" % ",".join(sum_fields))

        # Per Region values
        col = 3
        for region in region_headers:
            # Per Region values (sum of site values)
            for rownum, lic in enumerate(self.ordered_concurrent_lics, crow):
                region_sum_fields = []
                sites = self.mapper.region_sites(region)
                if region == 'Unknown':
                    sites = ['Unknown']
                for site in sites:
                    try:
                        target = lic_references[lic][site]
                    except KeyError:
                        pass
                    else:
                        key = "'%s'!%s" % (sheet_name, target)
                        region_sum_fields.append(key)
                if region_sum_fields:
                    self.set_cell(ws, rownum, col,
                                  "=SUM(%s)" % ",".join(region_sum_fields))
            col += 1
        dim = ws.column_dimensions['A']
        dim.width = 30.0

    def format_users(self, ws, sessions):
        self.log.debug("Creating Users table")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
        self.set_cell(ws, 1, 1, _(u'User Site Mapping'), 'PAGEHEADING')
        crow = 4

        headings = (u'Personalnummer',
                    u'Standort',
                    u'Zone',
                    u'Sitzungen')

        for col, heading in enumerate(headings, 1):
            self.set_cell(ws, crow, col, heading, 'HEADING')

        crow += 1
        # enumerate the users in our sessions
        cnt = Counter(s.uname for s in sessions)
        for user, count in cnt.most_common():
            site = self.mapper.map_user(user)
            self.set_cell(ws, crow, 1, user)
            self.set_cell(ws, crow, 2, site[1])
            self.set_cell(ws, crow, 3, site[2])
            self.set_cell(ws, crow, 4, count)

            if site[0] is None:
                # mark row in red
                for row in ws['A{0}:D{0}'.format(crow)]:
                    for c in row:
                        c.style = 'RED_WARN'
            crow += 1

    def format_failed_allocs(self, ws, sessions):
        self.log.debug("Creating Failed Allocs Table")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
        self.set_cell(ws, 1, 1, _(u'Failed License Allocations'), 'PAGEHEADING')
        crow = 4

        # Allocation frequency
        headings = (u'Lizenz',
                    u'Anzahl')
        for col, heading in enumerate(headings, 1):
            self.set_cell(ws, crow, col, heading, 'HEADING')

        crow += 1
        failed_cnt = Counter(l.mno for l in chain.from_iterable(
            s.failed_licenses for s in sessions))
        for mno, count in failed_cnt.most_common():
            self.set_cell(ws, crow, 1, mno)
            self.set_cell(ws, crow, 2, count)
            crow += 1

        # Detailed list of failed allocations
        crow += 5

        headings = (u'Date',
                    u'Lizenz',
                    u'Personalnummer',
                    u'Standort',
                    u'Zone')

        for col, heading in enumerate(headings, 1):
            self.set_cell(ws, crow, col, heading, 'HEADING')

        crow += 1

        for s in sessions:
            if not s.failed_licenses:
                continue
            site = self.mapper.map_user(s.uname)
            for lic in s.failed_licenses:
                self.set_cell(ws, crow, 1, lic.start_time)
                self.set_cell(ws, crow, 2, lic.mno)
                self.set_cell(ws, crow, 3, s.uname)
                self.set_cell(ws, crow, 4, site[1])
                self.set_cell(ws, crow, 5, site[2])
                crow += 1

    def format_session_info_data(self, ws, session_data, lbl_col, lbl_row, data_col, data_row):
        # TODO: use scientific libs instead of own code? matplotlib / numpy

        # generate bins for histogram
        bins = deque([x * 0.5 for x in range(1, 25)])

        bin_bounds = [0.0]
        bin_bound_max = 0.0
        bin_str = deque(['<' + str(bins[0])])

        # lower bound description
        last_el = '>' + str(bins[-1])

        # generate bound descriptions
        while len(bins) > 0:
            bin_bound_low = bins.popleft()
            bin_bound_max = bins.popleft()
            bin_bounds.append(bin_bound_low * 60 * 60)
            bin_str.append('%s - %s' % (bin_bound_low, bin_bound_max))

        # upper bound description
        bin_str.append(last_el)
        bin_bounds.append(bin_bound_max * 60 * 60)

        bin_counts = {}
        for i, dummy in enumerate(bin_str):
            bin_counts[i] = 0

        total = 0
        # count occurences of each bin
        for sess_dur in session_data:
            for i, bound in izip(reversed(xrange(len(bin_bounds))), reversed(bin_bounds)):
                if sess_dur > bound:
                    bin_counts[i] += 1
                    total += 1
                    break

        # generate tables
        c_desc = ws.cell(column=lbl_col, row=lbl_row)
        c_data = ws.cell(column=data_col, row=data_row)
        c_percent = ws.cell(column=data_col + 1, row=data_row)
        c_desc.value = _(u'Session Duration in Hours')
        c_data.value = _(u'Session Count')
        c_percent.value = _(u'Session Percent')
        c_desc.style, c_data.style = u'BORDERBR', u'BORDERB'
        c_percent.style = u"BORDERBR"
        ws.column_dimensions[get_column_letter(lbl_col)].width = 24
        ws.column_dimensions[get_column_letter(data_col)].width = 17
        ws.column_dimensions[get_column_letter(data_col + 1)].width = 20

        i_desc, i_data = lbl_row, data_row
        for bin_desc, bin_idx in zip(bin_str, bin_counts):
            i_desc += 1
            i_data += 1
            desc_cell = ws.cell(column=lbl_col, row=i_desc)
            desc_cell.value = bin_desc
            desc_cell.style = u'BORDERR_CENTERED'
            data_cell = ws.cell(column=data_col, row=i_data)
            data_cell.value = bin_counts[bin_idx]
            data_cell.style = u'CENTERED'
            percent_cell = ws.cell(column=data_col + 1, row=i_data)
            if total > 0:
                percent_cell.value = bin_counts[bin_idx] / (1.0 * total)
            percent_cell.style = u'CENTEREDPCT'

        i_data += 1
        cell = ws.cell(column=lbl_col, row=i_data)
        cell.value = _(u'Total')
        cell.style = u'BORDERR_CENTERED'
        cell = ws.cell(column=data_col, row=i_data)
        cell.value = total
        cell.style = u'CENTERED'

        self.layout_session_info_chart(ws, lbl_row + 1, lbl_col, data_row + 1, data_col,
                                       u'M6', c_desc.value, c_data.value)

    @staticmethod
    def layout_session_info_chart(ws, lbl_row, lbl_col, data_row, data_col, anchor, x_title, y_title):
        histogram = BarChart()
        histogram.type = u'col'
        histogram.style = 18
        histogram.width = 20
        histogram.height = 10
        histogram.title = _(u'Session Distribution')
        histogram.x_axis.title = x_title
        histogram.y_axis.title = y_title
        labels = Reference(ws, min_col=lbl_col, min_row=lbl_row, max_row=lbl_row + 13)
        data = Reference(ws, min_col=data_col, min_row=data_row - 1, max_row=data_row + 13)
        histogram.add_data(data, titles_from_data=True)
        histogram.set_categories(labels)
        ws.add_chart(histogram, anchor)

    @staticmethod
    def add_legend(ws, legend_entries, col_start, row_start):
        leg_start = row_start
        # Create legend entries
        for k, v in legend_entries.items():
            desc_cell = ws.cell(row=row_start, column=col_start)
            desc_cell.value = k
            desc_cell.border = Border(right=Side(style='medium'))
            val_cell = ws.cell(row=row_start, column=col_start + 1)
            val_cell.value = v
            row_start += 1
        # Add style to legend heading
        ws.cell(row=leg_start, column=col_start).border = Border(bottom=Side(style='medium'),
                                                                 right=Side(style='medium'))
        ws.cell(row=leg_start, column=col_start + 1).border = Border(bottom=Side(style='medium'))

    def format_session_info(self, ws, sessions):
        self.log.debug("Creating session table")
        ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=5)
        self.set_cell(ws, 1, 1, _(u'License Sessions'), 'PAGEHEADING')
        row_startnum = 6

        headings = (_(u'Session ID'),
                    _(u'Start Time'),
                    _(u'End Time'),
                    _(u'Duration'),
                    _(u'Site'),
                    _(u'User'),
                    _(u'Licenses'))

        # var to store histogram data
        session_data = []
        MAXROW = 1048576
        BASE_CROW = row_startnum
        COL_OFFSET_STEP = 20
        col_offset = -COL_OFFSET_STEP
        crow = row_startnum

        # dump the data
        for s in sessions:
            if not s.included_licenses and s.failed_licenses:
                # skip failed sessions
                continue
            if crow == MAXROW - 1 or crow == BASE_CROW:
                # open a new column set and write a header set
                crow = BASE_CROW
                col_offset += COL_OFFSET_STEP
                for col, heading in enumerate(headings, 1):
                    self.set_cell(ws, crow, col + col_offset, heading, 'HEADING')

                ws.column_dimensions.group(
                    get_column_letter(col_offset + 1),
                    hidden=True)

                dim = ws.column_dimensions[get_column_letter(col_offset + 2)]
                dim.width = 18.0
                dim = ws.column_dimensions[get_column_letter(col_offset + 3)]
                dim.width = 18.0

            crow += 1
            data = (s.pid,
                    s.dt_start(),
                    s.dt_end(),
                    s.duration(),
                    self.mapper.map_user(s.uname)[1],
                    s.uname,
                    ",".join(("%s[%s]" % (lic.mno, lic.lkind)
                              for lic in sorted(s.included_licenses,
                                                key=lambda x: x.mno))))
            session_data.append(s.duration().total_seconds())
            for col, item in enumerate(data, 1):
                c = ws.cell(row=crow, column=col + col_offset)
                c.value = item
                if col in (2, 3):
                    c.style = 'TIMESTAMP_DE'
                elif col == 4:
                    c.style = 'DURATION'
                elif col == 5 and item == u'Unknown':
                    # colorize red, warning
                    c.style = 'RED_WARN'

        if col_offset > 0:
            filter_crow = MAXROW
        else:
            filter_crow = crow
        # Set filter on data area of session worksheet (Bi:Gj)
        ws.auto_filter.ref = 'B%s:G%s' % (row_startnum, filter_crow)

        # Add legend for license types at top area
        # Entries to display
        legend_entries = OrderedDict([
            (_(u'License Type'), _(u'Description')),
            (u'SD', u'SOED'),
            (u'FX', u'Named User Paket'),
            (u'FL', u'Floating Session / Floating Option'),
            (u'EX', u'Integration'),
        ])
        # generate legend
        self.add_legend(ws, legend_entries, 9, 1)

        # Add histogram to worksheet
        self.format_session_info_data(ws, session_data, 13, 26, 14, 26)

    def render_overview(self, ws, dataset):
        header_end_row = self.format_header(ws, dataset)
        self.format_basic_usage_metrics(ws, header_end_row + 1, dataset)

    def new_sheet(self, title):
        # max length for titles is 31 Chars.
        self.log.debug("Creating new work sheet %s", title)
        if len(self.wb.sheetnames) == 1:
            sheet = self.wb.active
            if not sheet.title == "Sheet":
                sheet = self.wb.create_sheet()
        else:
            sheet = self.wb.create_sheet()
        sheet.title = title
        return sheet

    def render_sheet(self, title, func, *args):
        ws = self.new_sheet(title)
        return func(ws, *args)

    def patch_sheet(self, title, func, *args):
        ws = self.wb[title]
        return func(ws, *args)

    def format_report(self, dataset, sessions, start_time, end_time):
        self.start = utils.parse_timestamp(start_time)
        self.end = utils.parse_timestamp(end_time)

        first_sheet = u"Nutzer und Standorte"
        self.render_sheet(first_sheet,
                          self.render_overview,
                          dataset)

        self.render_sheet(u"Lizenzbedarf (Standortbezogen)",
                          self.format_daily_concurrent_sessions,
                          sessions)

        # we need back references to this sheet
        sheet_name = u"Lizenzbedarf (Summe Standorte)"
        lic_max_references = {}
        self.render_sheet(sheet_name,
                          self.format_daily_concurrent_sessions_max,
                          sessions,
                          lic_max_references)

        # Now patch the max values into the first page
        self.patch_sheet(first_sheet,
                         self.patch_max_per_site,
                         lic_max_references,
                         sheet_name)

        if self.mapper.have_regions():
            self.render_sheet(u"Lizenzbedarf (Summe Zonen)",
                              self.format_daily_concurrent_sessions_reg_max,
                              sessions,
                              lic_max_references,
                              sheet_name)

        self.render_sheet(u"Startzeiten",
                          self.format_start_time_daily,
                          dataset)

        self.render_sheet(u"Tagesnutzung",
                          self.format_mod_usage_daily,
                          dataset)

        self.render_sheet(u"Host Usage",
                          self.format_hosts,
                          dataset)

        self.render_sheet(u"Sitzungen (Dauer, Pakete)",
                          self.format_session_info,
                          sessions)

        self.render_sheet(u"Lizenzpakete (Standorte)",
                          self.format_license_sets_site,
                          sessions)

        self.render_sheet(u"Lizenzpakete (Anzahl)",
                          self.format_license_sets_user,
                          sessions)

        self.render_sheet(u"Lizenznamen",
                          self.format_license_info,
                          dataset)

        self.render_sheet(u"Nutzerstandorte",
                          self.format_users,
                          sessions)

        self.render_sheet(u"Fehlgeschlagene Allokationen",
                          self.format_failed_allocs,
                          sessions)

    def write_report(self, filename):
        self.log.info("Writing workbook to %s", filename)
        while 1:
            try:
                self.wb.save(filename)
            except IOError as exc:
                print(exc)
                ok = raw_input("Retry (y/n)?")
                if ok == 'y':
                    continue
            break

        self.log.debug("Wrote workbook")


def persist_sessions(filename, sessions):
    """Dump the session information as json"""
    log = logging.getLogger('cdblic.persist')
    log.info("Persisting %d sessions to %s", len(sessions), filename)

    d = []
    for sess in sessions:
        stream_id = 0 if not hasattr(sess, 'stream_id') else sess.stream_id
        lics = [(l.mno, l.start_time, l.lkind) for l in sess.included_licenses]
        flics = [(l.mno, l.start_time, l.lkind) for l in sess.failed_licenses]
        obj = [sess.pid, stream_id, sess.uname, sess.start_time, sess.end_time, lics, flics]
        d.append(obj)
    with open(filename, 'wb') as fd:
        json.dump(d, codecs.getwriter('utf-8')(fd), indent=2, ensure_ascii=False)


def parse_sessions(filename):
    """Load a persisted session file"""
    with open(filename, 'rb') as fd:
        obj = json.load(fd)

    sessions = []
    # share uname objects instances
    _uname = {}
    # share licinfos objects
    _mno = {}
    _lkind = {}

    li = lsession.LicInfo
    for item in obj:
        s = lsession.Session()
        s.pid, s.stream_id, uname, s.start_time, s.end_time, lics, failed = item
        s.uname = _uname.setdefault(uname, uname)
        s.included_licenses = [li(_mno.setdefault(mno, mno), start_time, _lkind.setdefault(lkind, lkind))
                               for (mno, start_time, lkind) in lics]
        s.failed_licenses = [li(_mno.setdefault(mno, mno), start_time, _lkind.setdefault(lkind, lkind))
                             for (mno, start_time, lkind) in failed]
        sessions.append(s)

    # Reduce memory footprint
    obj = None
    _uname = None
    _mno = None
    _lkind = None
    gc.collect()
    return sessions


def get_min_max_starttimes(tables):
    rset = sqlapi.RecordSet2(
        table=tables['lstatistics'],
        sql="SELECT MIN(lbtime) as mintime, MAX(lbtime) AS maxtime FROM %(lstatistics)s" % tables)
    timetuple = (rset[0]['mintime'], rset[0]['maxtime'])
    return timetuple


class Identity(object):

    def get_id(self, prefix, name):
        return name

    def __call__(self, event):
        return event


def generate_report(versionstr='', customer='',
                    start_time=None, end_time=None,
                    reportfile=None,
                    mode='session',
                    do_extract=True,
                    do_report=True,
                    tables=None,
                    pseudonymize=True):

    log = logging.getLogger('cdblic.generate')

    if not versionstr and do_extract:
        versionstr = lsession.get_cdb_version(tables['lstatistics'])
        # major, minor = swversion()
        # don't use version of software since table data may be imported from another installation
        # versionstr = "%s.%s" % (major, minor)

    if not (start_time and end_time):
        timetuple = get_min_max_starttimes(tables)
        if not start_time:
            start_time = timetuple[0]
        if not end_time:
            end_time = timetuple[1]

    # Identity default mapper
    event_mapper = Identity()
    if pseudonymize:
        event_mapper = lsession.EventPseudonymizer()

    fmode = mode
    site_mapper = None
    statsdata = None

    if reportfile is None:
        reportfilepath = os.path.join(
            os.environ["CADDOK_TMPDIR"],
            'LicReport_%s_%s.xlsx' % (customer, fmode))
    else:
        reportfilepath = os.path.join(
            os.environ["CADDOK_TMPDIR"],
            reportfile)

    statsfile = os.path.splitext(reportfilepath)[0] + u'.stats.json'
    psfile = os.path.splitext(reportfilepath)[0] + u'.sessions.json'
    sitefile = os.path.splitext(reportfilepath)[0] + u'.sitemap.json'

    if do_extract:
        # Pseudonymize during extraction
        site_mapper = lsite_map.SiteMapper(tables, event_mapper)
        coll = Collector(versionstr, customer, start_time, end_time, tables,
                         pseudonymizer=event_mapper)
        coll.collect()

        # Dump Generic Statistics
        log.info("Persisting %d statistics to %s", len(coll.data), statsfile)
        with open(statsfile, 'wb') as f:
            coll.dump(f)
        statsdata = coll.data

        extractor = lsession.extract_sessions

        all_sessions = extractor(start_time, end_time, tables,
                                 mapper=event_mapper)
        persist_sessions(psfile, all_sessions)

        log.info("Persisting site info to %s", sitefile)
        with open(sitefile, 'wb') as f:
            site_mapper.dump(f)

    site_mapper = None
    statsdata = None
    all_sessions = None

    if do_report:
        # Reporting, load data or use the already loaded dataset
        event_mapper = Identity()
        if not site_mapper:
            log.info("Loading site info from %s", sitefile)
            site_mapper = lsite_map.SiteMapper(tables, event_mapper)
            with open(sitefile, 'rb') as f:
                site_mapper.load(f)
        if not statsdata:
            log.info("Loading stats info from %s", statsfile)
            with open(statsfile, 'rb') as f:
                statsdata = json.load(f)
        if not all_sessions:
            log.info("Loading sessions from %s", psfile)
            all_sessions = parse_sessions(psfile)

        # Reduce memory footprint
        gc.collect()

        start_time = statsdata['Basic.start_time']
        end_time = statsdata['Basic.end_time']
        versionstr = statsdata['Basic.cdb_version']
        version = int(versionstr.split('.')[0])
        formatter = XLSXFormatter(site_mapper, version)
        formatter.format_report(statsdata,
                                all_sessions,
                                start_time,
                                end_time,
                                )
        # Reduce memory footprint
        all_sessions = None
        statsdata = None
        site_mapper = None
        gc.collect()
        formatter.write_report(reportfilepath)
    return reportfilepath


def swversion():
    from cdb import version
    swversion = version.verstring(1)
    major, minor = map(str, swversion.split('.')[:2])
    return major, minor
