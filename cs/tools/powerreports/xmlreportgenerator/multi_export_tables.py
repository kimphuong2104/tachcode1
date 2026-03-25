from copy import copy

from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.table import Table

from cs.tools.powerreports.xmlreportgenerator import tools
from cs.tools.powerreports.xmlreportgenerator.worksheet_tools import WorksheetTools


class MultiExports(object):
    def __init__(self, worksheet):
        self.worksheet = worksheet

    def move_row_height(self, start_row):
        idx = 0
        row_size = []
        for row in range(1, self.worksheet.max_row + start_row):
            size = self.worksheet.row_dimensions[row].height
            row_size.append(size)
        for row in range(1 + start_row, self.worksheet.max_row + start_row):
            self.worksheet.row_dimensions[row].height = row_size[idx]
            idx += 1

    def move_cdb_template_input(self, export_values):
        start_row = export_values[0]
        start_column = export_values[1]

        # self.worksheet.merged_cells.ranges is set
        # change to list for get ranges as string to unmerge cells

        merged_cells = [
            str(merged_cell_range)
            for merged_cell_range in self.worksheet.merged_cells.ranges
        ]
        for merged_cell_range in merged_cells:
            self.worksheet.unmerge_cells(merged_cell_range)
        if start_row > 0:
            self.worksheet.insert_rows(0, start_row)
            self.move_row_height(start_row)
        if start_column > 0:
            self.worksheet.insert_cols(0, start_column)

        if len(self.worksheet.tables) != 0:
            for sheet_table in self.worksheet.tables.values():
                table_size = sheet_table.ref.split(":")
                table_start_col, table_start_row = tools.letter_num_seperated(
                    table_size[0]
                )
                table_end_col, table_end_row = tools.letter_num_seperated(table_size[1])
                table_start_col = (
                    column_index_from_string(table_start_col) + start_column
                )
                table_start_col = get_column_letter(table_start_col)
                table_start_row = str(int(table_start_row) + start_row)
                table_end_col = column_index_from_string(table_end_col) + start_column
                table_end_col = get_column_letter(table_end_col)
                table_end_row = str(int(table_end_row) + start_row)

                sheet_table.ref = (
                    table_start_col
                    + table_start_row
                    + ":"
                    + table_end_col
                    + table_end_row
                )

    def copy_template_target(self, cdb_ws, start, end):
        """
        Copy everything input from template sheet to target worksheet
        :param cdb_ws: template sheet
        :param string start: start of template sheet
        :param str end: end of template sheet (used ranged/max row oder filled table)

        Change worksheet column width like cdb_ws

        Put temp cell (eg A12) value in sheet cell (eg A12)
        Copy cell style (fill, alignment etc) from temp to sheet
        Remove first row of worksheet, if first row has value target!$xx
        """
        range_ws = str(start) + ":" + str(end)
        for row in cdb_ws[range_ws]:
            for cell in row:
                self.worksheet[cell.coordinate].value = cell.value
                self.worksheet.row_dimensions[cell.row].height = cdb_ws.row_dimensions[
                    cell.row
                ].height
                self.worksheet[cell.coordinate].alignment = copy(cell.alignment)
                self.worksheet[cell.coordinate].fill = copy(cell.fill)
                self.worksheet[cell.coordinate].font = copy(cell.font)
                self.worksheet[cell.coordinate].border = copy(cell.border)
                self.worksheet[cell.coordinate].number_format = copy(cell.number_format)
                self.worksheet[cell.coordinate].protection = copy(cell.protection)

    def add_formats_to_sheet(
        self,
        export_start,
        export_offsets,
        hidden_rows_columns,
        single_cell_space,
        tables_infos,
        merged_cells_in_ws,
    ):
        """
        Adding formats (tables, merged_cells, hyperlinks) to sheet
        :param string export_start: start of template sheet
        :param list export_offsets: list of offsets for every export
        :param dict hidden_rows_columns: position of hidden rows and hidden columns
        :param int single_cell_space: space of single cells to table_start
        :param dict tables_infos: for every table key 1-... with start, end and style
        :param list merged_cells_in_ws: ranges of merged_cell from template


        Reverse tables_infos and export_offsets to start with first table/export

        Every key/table in tables_info == 1 export named idx

        For each export get export_offset from list
        Get ranges of merged_cells of first export (was set already)
        Shift cel range by export_offset - first export_offset (if export has a start at eg A6)
        Merge shifted ranges

        Get table size from original table start/end in tables_infos
        Use idx offset and single_cell_spaces
        Next offset is changed beforehand with current offset + next offset
        Name table with idx individually
        """

        tables_infos = tools.get_dict_reversed(tables_infos)
        export_offsets.reverse()
        table_start, table_end = None, None

        for idx, table_infos in tables_infos.items():
            export_offset = export_offsets[idx]
            for cell_range in merged_cells_in_ws:
                cell_range = CellRange(cell_range)
                cell_range.shift(row_shift=export_offset - export_offsets[0])
                self.worksheet.merge_cells(str(cell_range))

            start_offset = export_offset + single_cell_space
            table_start, table_end = tools.get_start_end_of_table(
                [table_infos["Start"], table_infos["End"]]
            )
            table_size = int(table_end[1]) - int(table_start[1])
            if export_start[1:] != "1":
                start_offset = start_offset - 1

            table_start = table_start[0] + str(start_offset)
            table_end = table_end[0] + str(start_offset + table_size)
            export_offsets[idx + 1] = export_offset + export_offsets[idx + 1]

            for row in hidden_rows_columns["Rows"]:
                self.worksheet.row_dimensions[
                    row + export_offset - export_offsets[0]
                ].hidden = True
            for column in hidden_rows_columns["Columns"]:
                self.worksheet.column_dimensions[
                    get_column_letter(column)
                ].hidden = True
            if idx > 0:
                row_id = 0
                for row in range(
                    start_offset, len(table_infos["RowSizes"]) + start_offset
                ):
                    self.worksheet.row_dimensions[row].height = table_infos["RowSizes"][
                        row_id
                    ]
                    row_id += 1

            if table_infos["Style"] is not None:
                if table_infos["Style"] == "Tables":
                    self.multi_tables(
                        table_infos["Tables"],
                        start_offset,
                        export_start,
                        idx,
                        hidden_rows_columns,
                    )
                else:
                    tab = Table(
                        displayName="Tabelle1" + str(idx + 1) + str(idx),
                        ref=table_start + ":" + table_end,
                    )
                    tab.tableStyleInfo = table_infos["Style"]
                    self.worksheet.add_table(tab)

        sheet_range = export_start + ":" + table_end
        self.add_hyperlinks(sheet_range)

    def multi_tables(
        self, table_infos, start_offset, export_start, idx, hidden_table_rows
    ):
        for table_id, table_info in table_infos.items():
            if idx >= 1:
                export_value = (
                    int(export_start[1:]) + start_offset - (int(export_start[1:]) * 2)
                )
                hidden_table_rows["Table"][:] = [
                    i + export_value for i in hidden_table_rows["Table"]
                ]
            else:
                export_value = int(export_start[1:]) - start_offset
            table_start, table_end = tools.get_start_end_of_table(
                [table_info["Start"], table_info["End"]]
            )
            table_start[1] = int(table_start[1]) + export_value
            table_end[1] = int(table_end[1]) + export_value

            for row_id, row in enumerate(hidden_table_rows["Table"]):
                if (
                    row_id == table_id
                    and self.worksheet.row_dimensions[row].hidden is True
                    and row != table_start[1]
                ):
                    self.worksheet.row_dimensions[row].hidden = False
            hidden_table_rows["Table"][:] = [
                i - export_value for i in hidden_table_rows["Table"]
            ]
            table_start = table_start[0] + str(table_start[1])
            table_end = table_end[0] + str(table_end[1])

            tab = Table(
                displayName="Tabelle1" + str(idx + 1) + str(table_id),
                ref=table_start + ":" + table_end,
            )
            tab.tableStyleInfo = table_info["Style"]
            self.worksheet.add_table(tab)

    def add_hyperlinks(self, sheet_range):
        """
        Adding hyperlinks in multi-exports

        Using WorksheetTools method
        Value of cell, which should have a hyperlink, has ' cdb:texttodisplay:' in text
        """
        for row in self.worksheet[sheet_range]:
            for cell in row:
                if type(cell.value) is str and " cdb:texttodisplay:" in cell.value:
                    WorksheetTools(self.worksheet).create_cell_hyperlink(
                        cell.value, cell.coordinate
                    )
