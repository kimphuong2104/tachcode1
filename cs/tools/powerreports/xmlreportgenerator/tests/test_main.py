# -*- coding: utf-8 -*-

import datetime
import os
import shutil
import tempfile
import zipfile

import openpyxl as xl
from openpyxl.utils.cell import get_column_letter

from cdb import testcase
from cs.tools.powerreports.xmlreportgenerator import ExcelReportGenerator, tools
from cs.tools.powerreports.xmlreportgenerator.handle_xml import XMLFile
from cs.tools.powerreports.xmlreportgenerator.single_cell_importer import (
    SingleCellsImporter,
)
from cs.tools.powerreports.xmlreportgenerator.table_importer import TableImporter

WORKING_DIR = os.path.dirname(__file__)
TEST_DIR = os.path.join(
    WORKING_DIR, "..", "..", "..", "..", "..", "tests", "test_files"
)


class ExcelGeneratorTestCase(testcase.RollbackTestCase):
    def setUp(self):
        super(ExcelGeneratorTestCase, self).setUp()
        self.test_dir = tempfile.mkdtemp()
        self.cdbxml_dir = tempfile.mkdtemp()
        with zipfile.ZipFile(
            os.path.join(TEST_DIR, "Anforderungsübersicht_data.xlsx.cdbxml.zip")
        ) as excel_file:
            excel_file.extractall(self.cdbxml_dir)
        self.excel_dir = tempfile.mkdtemp()
        with zipfile.ZipFile(
            os.path.join(TEST_DIR, "Anforderungsübersicht_template.xlsx")
        ) as excel_file:
            excel_file.extractall(self.excel_dir)
        shutil.copy2(
            os.path.join(TEST_DIR, "Anforderungsübersicht_template.xlsx"), self.test_dir
        )
        shutil.copy2(
            os.path.join(TEST_DIR, "ReportPersons_template.xlsx"), self.test_dir
        )
        shutil.copy2(
            os.path.join(TEST_DIR, "VarStücklistenvergleich_de.xlsm"), self.test_dir
        )
        shutil.copy2(
            os.path.join(TEST_DIR, "VarStücklistenvergleich_cs.variants_de.xlsm"),
            self.test_dir,
        )
        shutil.copy2(os.path.join(TEST_DIR, "Analysesteckbrief_de.xlsm"), self.test_dir)
        shutil.copy2(os.path.join(TEST_DIR, "Aufwendungen_proj_de.xlsx"), self.test_dir)
        shutil.copy2(
            os.path.join(TEST_DIR, "Mitarbeiterauslastung_de.xlsx"), self.test_dir
        )
        shutil.copy2(os.path.join(TEST_DIR, "PriceReport_en.xlsm"), self.test_dir)

    def test_insert_in_table(self):
        xml_data_file = os.path.join(
            self.cdbxml_dir, "Anforderungsübersicht_data.xlsx.cdbxml"
        )
        template = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")

        xml = XMLFile(xml_data_file)
        xml_data = xml.get_single_key("Root")
        table_file_path = os.path.join(self.excel_dir, "xl", "tables", "table1.xml")

        workbook = xl.load_workbook(template)

        sheet1 = workbook.worksheets[0]
        table = sheet1.tables.get("Tabelle6")
        table_import = TableImporter(
            sheet1, [""], table, xml_data, table_file_path, xml_data_file
        )
        table_import.insert_in_table()

        expected_new_table_size = "B27:L41"
        table_size = table.ref

        self.assertEqual(expected_new_table_size, table_size)
        expected_value_c29 = "Anf. mit Akzeptanz Kriteria"
        cell_c29 = sheet1["C29"].value
        self.assertEqual(expected_value_c29, cell_c29)

    def test_excel_arguments_prop(self):
        excel_custom_props = {
            "Arguments": [
                "cdbxml_report_date",
                "cdbxml_report_datetime",
                "cdbxml_report_author",
                "cdbxml_report_lang",
            ]
        }
        data = os.path.join(TEST_DIR, "Anforderungsübersicht_data.xlsx.cdbxml.zip")
        template = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")

        ExcelReportGenerator(template, data, excel_custom_props).generate()

        workbook = xl.load_workbook(template)
        expected_props = [
            "Arguments.cdbxml_report_date",
            "Arguments.cdbxml_report_datetime",
            "Arguments.cdbxml_report_author",
            "Arguments.cdbxml_report_lang",
        ]
        props = []
        for prop in workbook.custom_doc_props.props:
            props.append(prop.name)
        self.assertEqual(props, expected_props)

    def test_single_cells_import(self):
        xml_data_file = os.path.join(
            self.cdbxml_dir, "Anforderungsübersicht_data.xlsx.cdbxml"
        )
        template = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")

        xml = XMLFile(xml_data_file)
        xml_data = xml.get_single_key("Root")
        singl_cell_file_path = os.path.join(
            self.excel_dir, "xl", "tables", "tableSingleCells1.xml"
        )

        workbook = xl.load_workbook(template)

        sheet1 = workbook.worksheets[0]
        single_cell_import = SingleCellsImporter(
            workbook, {}, sheet1, singl_cell_file_path, xml_data, None, None
        )
        single_cell_import.insert_in_single_cells()

        expected_value_d4 = "Diff Plugins All Cases"
        cell_d4 = sheet1["D4"].value
        self.assertEqual(expected_value_d4, cell_d4)

        expected_datetime_i1 = "14.10.2022"
        expected_datetime_i1 = datetime.datetime.strptime(
            expected_datetime_i1, "%d.%m.%Y"
        )
        datetime_i1 = sheet1["I1"].value
        self.assertEqual(expected_datetime_i1, datetime_i1)

    def test_generate_excel_valueerror(self):
        data = os.path.join(
            TEST_DIR,
            "Anforderungsübersicht_Wed-03-May-2023-13-34-28_caddok.xlsx.cdbxml.zip",
        )
        template = os.path.join(
            TEST_DIR, "Anforderungsübersicht_Wed-03-May-2023-13-34-28_caddok.xlsx"
        )

        res = ExcelReportGenerator(template, data).generate()
        self.assertNotEqual(res["status"], "OK")

    def test_generate_excel(self):
        data = os.path.join(TEST_DIR, "Anforderungsübersicht_data.xlsx.cdbxml.zip")
        template = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")

        ExcelReportGenerator(template, data).generate()

        workbook = xl.load_workbook(template)
        sheet1 = workbook.worksheets[0]

        expected_value_c29 = "Anf. mit Akzeptanz Kriteria"
        cell_c29 = sheet1["C29"].value
        self.assertEqual(expected_value_c29, cell_c29)

        expected_value_d4 = "Diff Plugins All Cases"
        cell_d4 = sheet1["D4"].value
        self.assertEqual(expected_value_d4, cell_d4)

        sheet3 = workbook.worksheets[2]
        expected_value_a1 = "Bewertungen"
        cell_a1 = sheet3["A1"].value
        self.assertEqual(expected_value_a1, cell_a1)

    def test_generate_sum_but_no_total_row(self):
        data = os.path.join(
            TEST_DIR, "Aufwände_Thu-26-Jan-2023-09-57-46_caddok.xlsx.cdbxml.zip"
        )
        template = os.path.join(self.test_dir, "Aufwendungen_proj_de.xlsx")
        workbook = xl.load_workbook(template)
        sheet_o = workbook.worksheets[0]

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook.worksheets[0]

        self.assertEqual(sheet["F14"].value, sheet_o["F13"].value)
        self.assertEqual(sheet["G14"].value, "=SUBTOTAL(109,Tabelle8[Stunden])")

    def test_generate_total_row_and_not_only_change_table_size(self):
        data = os.path.join(
            TEST_DIR,
            "Varianten-Stücklistenvergleich_Tue-03-Jan-2023-12-21-22_caddok.xlsm.cdbxml.zip",
        )
        template = os.path.join(self.test_dir, "VarStücklistenvergleich_de.xlsm")
        workbook = xl.load_workbook(template)
        sheet_o = workbook.worksheets[0]

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook.worksheets[0]

        self.assertEqual(sheet["S13"].value, sheet_o["S13"].value)
        self.assertEqual(sheet["A189"].value, sheet_o["A12"].value)
        self.assertEqual(sheet["G189"].value, "=SUBTOTAL(109,Tabelle3[Variante 1])")
        self.assertEqual(sheet["F190"].value, sheet_o["F13"].value)
        self.assertEqual(sheet["F205"].value, sheet_o["F28"].value)

    def test_generate_variants_report_group(self):
        data = os.path.join(
            TEST_DIR,
            "Varianten-Stücklistenvergleich_Tue-06-Jun-2023-08-28-36_caddok.xlsm.cdbxml.zip",
        )
        template = os.path.join(
            self.test_dir, "VarStücklistenvergleich_cs.variants_de.xlsm"
        )
        workbook = xl.load_workbook(template)
        sheet_o = workbook.worksheets[0]

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook.worksheets[0]
        self.assertEqual(sheet["S13"].value, sheet_o["S13"].value)

    def test_generate_excel_cdb_image_main(self):
        data = os.path.join(
            TEST_DIR,
            "InnoAnalysisOverview_Fri-20-Jan-2023-09-56-23_caddok.xlsm.cdbxml.zip",
        )

        template = os.path.join(self.test_dir, "Analysesteckbrief_de.xlsm")
        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook.worksheets[0]

        worksheet_images = tools.sheet_images(sheet)
        # one hidden image + 2 images
        self.assertIs(len(worksheet_images), 2)

        for image in worksheet_images:
            if "cdb_image" in image.anchor.pic.nvPicPr.cNvPr.name:
                temp_image_is_hidden = image.anchor.pic.nvPicPr.cNvPr.hidden
                self.assertTrue(temp_image_is_hidden)
            else:
                image_column = get_column_letter(tools.anchor_from(image).col + 1)
                image_row = tools.anchor_from(image).row + 1
                image_anchor = image_column + str(image_row)
                self.assertEqual(image_anchor, "B5")

    def test_generate_excel_cdb_image_table(self):
        data = os.path.join(TEST_DIR, "ReportPersons_data.xlsx.cdbxml.zip")
        template = os.path.join(self.test_dir, "ReportPersons_template.xlsx")

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook.worksheets[0]

        worksheet_images = tools.sheet_images(sheet)
        # one hidden image + 2 images
        self.assertIs(len(worksheet_images), 3)

        for image in worksheet_images:
            if "cdb_image" in image.anchor.pic.nvPicPr.cNvPr.name:
                temp_image_is_hidden = image.anchor.pic.nvPicPr.cNvPr.hidden
                self.assertTrue(temp_image_is_hidden)
            else:
                image_column = tools.anchor_from(image).col + 1
                self.assertEqual(image_column, 2)

    def test_column_changed_labels(self):
        data = os.path.join(
            TEST_DIR,
            "Mitarbeiterauslastung_Wed-24-May-2023-16-14-11_caddok.xlsx.cdbxml.zip",
        )
        template = os.path.join(self.test_dir, "Mitarbeiterauslastung_de.xlsx")

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook.worksheets[0]

        self.assertEqual(sheet["B4"].value, sheet["B3"].value)
        self.assertEqual(sheet["C4"].value, sheet["C3"].value)

    def test_table_column_formula(self):
        data = os.path.join(
            TEST_DIR,
            "SalesPriceReport (importable).xlsm.cdbxml.zip",
        )
        template = os.path.join(self.test_dir, "PriceReport_en.xlsm")

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook.worksheets[1]

        # original formula =IF(ISBLANK([@[Sub Code]]);[@[Machine Code]];[@[Sub Code]])
        # openpyxl changes formula when put in column to work for table rows
        expected_formula = "=IF(ISBLANK(Table1[[#This Row],[Sub Code]])," \
                           "Table1[[#This Row],[Machine Code]]," \
                           "Table1[[#This Row],[Sub Code]])"

        # formula in first row of table
        self.assertEqual(expected_formula, sheet["H2"].value)

        # same formula in row below
        self.assertEqual(expected_formula, sheet["H3"].value)
        self.assertEqual(expected_formula, sheet["H4"].value)

        expected_formula2 = "=valid_to"

        # formula in first row of table
        self.assertEqual(expected_formula2, sheet["O2"].value)

        # same formula in row below
        self.assertEqual(expected_formula2, sheet["O3"].value)
        self.assertEqual(expected_formula2, sheet["O4"].value)

    def tearDown(self):
        super(ExcelGeneratorTestCase, self).tearDown()
        shutil.rmtree(self.test_dir)
        shutil.rmtree(self.cdbxml_dir)
        shutil.rmtree(self.excel_dir)
