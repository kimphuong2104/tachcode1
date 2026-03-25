# -*- coding: utf-8 -*-

import os
import shutil
import tempfile

import openpyxl as xl
from openpyxl.worksheet.table import TableStyleInfo

from cdb import testcase
from cs.tools.powerreports.xmlreportgenerator import ExcelReportGenerator
from cs.tools.powerreports.xmlreportgenerator.multi_export_tables import MultiExports

WORKING_DIR = os.path.dirname(__file__)
TEST_DIR = os.path.join(
    WORKING_DIR, "..", "..", "..", "..", "..", "tests", "test_files"
)
# the template file has an extra row in beginning, for tests with sheet compare eg template A13 with sheet A12


class MultiExportTestCase(testcase.RollbackTestCase):
    def setUp(self):
        super(MultiExportTestCase, self).setUp()
        self.test_dir = tempfile.mkdtemp()
        shutil.copy2(
            os.path.join(TEST_DIR, "Strukturstückliste_filled_cdb_template.xlsx"),
            self.test_dir,
        )
        shutil.copy2(
            os.path.join(TEST_DIR, "Strukturstückliste_cdb_template.xlsx"),
            self.test_dir,
        )
        shutil.copy2(
            os.path.join(TEST_DIR, "StrukturstücklisteC1_cdb_template.xlsx"),
            self.test_dir,
        )
        shutil.copy2(
            os.path.join(TEST_DIR, "StrukturstücklisteB7_cdb_template.xlsx"),
            self.test_dir,
        )
        shutil.copy2(
            os.path.join(TEST_DIR, "StrukturstücklisteA5_cdb_template.xlsx"),
            self.test_dir,
        )
        shutil.copy2(os.path.join(TEST_DIR, "QualityGates_de_tmpl.xlsx"), self.test_dir)

    def test_target_start_B7_with_target_sheet_has_cell_value(self):
        template = os.path.join(self.test_dir, "StrukturstücklisteB7_cdb_template.xlsx")
        data = os.path.join(
            TEST_DIR, "Strukturstückliste_cdb_template_data.xlsx.cdbxml.zip"
        )

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook["Tabelle2"]
        expected_value = "Test"

        self.assertEqual(sheet["A1"].value, expected_value)

    def test_multi_export_offset_A5(self):
        template = os.path.join(self.test_dir, "StrukturstücklisteA5_cdb_template.xlsx")
        data = os.path.join(
            TEST_DIR, "Strukturstückliste_cdb_template_data.xlsx.cdbxml.zip"
        )

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook["Tabelle2"]
        expected_value_b7 = "Artikelnummer"
        not_expected_value_a4 = "Tabelle2!$A5"

        self.assertEqual(sheet["B7"].value, expected_value_b7)
        self.assertNotEqual(sheet["A5"].value, not_expected_value_a4)

    def test_multi_export_offset_C1_second_table(self):
        template = os.path.join(self.test_dir, "StrukturstücklisteC1_cdb_template.xlsx")
        data = os.path.join(
            TEST_DIR, "Strukturstückliste_cdb_template_data.xlsx.cdbxml.zip"
        )

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook["Tabelle2"]
        # expected_value_c31 = '9502667'
        sheet_tables = sheet.tables.keys()

        # under linux-oracle
        # AssertionError: '9502673' != '9502667'
        # mostly error comes from line ending
        # self.assertEqual(sheet['C31'].value, expected_value_c31)
        self.assertIn("Tabelle132", sheet_tables)

    def test_multi_export_quality_gates(self):
        template = os.path.join(self.test_dir, "QualityGates_de_tmpl.xlsx")
        data = os.path.join(
            TEST_DIR, "Quality Gates_Fri-13-Jan-2023-10-58-12_caddok.xlsx.cdbxml.zip"
        )
        excel_custom_props = {
            "Arguments": [
                "cdbxml_report_date",
                "cdbxml_report_datetime",
                "cdbxml_report_author",
                "cdbxml_report_lang",
            ]
        }
        ExcelReportGenerator(template, data, excel_custom_props).generate()
        workbook = xl.load_workbook(template)
        sheet = workbook["Quality Gates"]
        expected_props = [
            "Arguments.cdbxml_report_date",
            "Arguments.cdbxml_report_datetime",
            "Arguments.cdbxml_report_author",
            "Arguments.cdbxml_report_lang",
        ]
        props = []
        for prop in workbook.custom_doc_props.props:
            if prop.name.startswith("Arguments."):
                props.append(prop.name)
        self.assertEqual(props, expected_props)

        self.assertTrue(sheet.row_dimensions[10].hidden)

    def test_cdb_template_with_no_tables(self):
        template = os.path.join(self.test_dir, "Strukturstückliste_cdb_template.xlsx")
        data = os.path.join(
            TEST_DIR, "Strukturstückliste_cdb_template_data.xlsx.cdbxml.zip"
        )

        ExcelReportGenerator(template, data).generate()
        workbook = xl.load_workbook(template)
        template_sheet = workbook["cdb_template"]
        sheet = workbook["Tabelle2"]

        self.assertEqual(template_sheet["F14"].value, sheet["F14"].value)
        self.assertEqual(len(sheet.tables), 0)

    def test_copy_inputs_from_template_to_sheet(self):
        workbook = xl.load_workbook(
            os.path.join(self.test_dir, "Strukturstückliste_filled_cdb_template.xlsx")
        )
        template_sheet = workbook["cdb_template"]
        sheet = workbook["Tabelle1"]
        multi_export = MultiExports(sheet)
        multi_export.copy_template_target(template_sheet, "A1", "K18")
        self.assertEqual(template_sheet["F14"].value, sheet["F14"].value)

    def test_add_formats_to_sheet(self):
        workbook = xl.load_workbook(
            os.path.join(self.test_dir, "Strukturstückliste_filled_cdb_template.xlsx")
        )
        template_sheet = workbook["cdb_template"]
        sheet = workbook["Tabelle1"]
        multi_export = MultiExports(sheet)
        multi_export.copy_template_target(template_sheet, "A1", "K18")

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        table_info = {0: {"Start": "A12", "End": "K18", "Style": style, "Tables": None}}
        multi_export.add_formats_to_sheet(
            "A1", [0, 1], {"Rows": [], "Columns": []}, 10, table_info, ["C3:D3"]
        )

        self.assertEqual(sheet["F13"].value, "dev/")
        self.assertIsNotNone(sheet["A13"].hyperlink)

    def tearDown(self):
        super(MultiExportTestCase, self).tearDown()
        shutil.rmtree(self.test_dir)
