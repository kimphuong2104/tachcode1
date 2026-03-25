# -*- coding: utf-8 -*-

import os
from datetime import datetime

import openpyxl as xl

from cdb import testcase
from cs.tools.powerreports.xmlreportgenerator.worksheet_tools import WorksheetTools

WORKING_DIR = os.path.dirname(__file__)
TEST_DIR = os.path.join(
    WORKING_DIR, "..", "..", "..", "..", "..", "tests", "test_files"
)
EXCEL_FILE = os.path.join(TEST_DIR, "Anforderungsübersicht_template.xlsx")


class WorksheetToolsTestCase(testcase.RollbackTestCase):
    def test_add_hyperlink(self):
        workbook = xl.load_workbook(EXCEL_FILE)
        sheet = workbook["Tabelle1"]
        cdbxml_hyperlink = (
            "cdb:///byname/classname/cdbrqm_spec_object/CDB_ShowObject"
            " cdb:texttodisplay:RT000000103"
        )
        hyperlink_cell = "C34"
        worksheet_tools = WorksheetTools(sheet)
        worksheet_tools.create_cell_hyperlink(cdbxml_hyperlink, hyperlink_cell)

        self.assertEqual(sheet[hyperlink_cell].value, "RT000000103")
        self.assertIsNotNone(sheet[hyperlink_cell].hyperlink)

    def test_add_datetime(self):
        workbook = xl.load_workbook(EXCEL_FILE)
        sheet = workbook["Tabelle1"]
        cdbxml_type = "date"
        date_cell = "I1"
        date = "2022-10-14"
        worksheet_tools = WorksheetTools(sheet)
        worksheet_tools.set_cell_value(date, cdbxml_type, date_cell)

        self.assertEqual(type(sheet[date_cell].value), datetime)

    def test_copy_style_row(self):
        table_size = ["B27", "L30"]
        workbook = xl.load_workbook(EXCEL_FILE)
        sheet = workbook["Tabelle1"]
        worksheet_tools = WorksheetTools(sheet)
        for table in sheet.tables.values():
            worksheet_tools.copy_style_from_first_row(table, table_size)
        expected_alignment = "right"

        self.assertEqual(sheet["D30"].alignment.horizontal, expected_alignment)
        self.assertIs(sheet["C29"].alignment.wrapText, True)

    def test_extend_data_validation(self):
        workbook = xl.load_workbook(EXCEL_FILE)
        sheet = workbook["Tabelle1"]
        worksheet_tools = WorksheetTools(sheet)
        table_start = ["B", "27"]
        table_end = ["L", "28"]
        expected_data_validation = "K28:K31"
        current_data_validation = sheet.data_validations.dataValidation

        worksheet_tools.extend_validations_by_table_length(
            current_data_validation, table_start, table_end, 4
        )

        extended_data_validation = sheet.data_validations.dataValidation
        extended_data_validation = extended_data_validation[0]

        self.assertEqual(str(extended_data_validation.sqref), expected_data_validation)
