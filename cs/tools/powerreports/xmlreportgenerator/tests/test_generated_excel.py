# -*- coding: utf-8 -*-

import os
import shutil
import tempfile
from datetime import datetime

import openpyxl as xl
from openpyxl.utils.cell import get_column_letter
from PIL import ImageColor

from cdb import testcase
from cs.tools.powerreports.xmlreportgenerator import ExcelReportGenerator

WORKING_DIR = os.path.dirname(__file__)
TEST_DIR = os.path.join(
    WORKING_DIR, "..", "..", "..", "..", "..", "tests", "test_files"
)


class GeneratedExcelTestCase(testcase.RollbackTestCase):
    def setUp(self):
        super(GeneratedExcelTestCase, self).setUp()
        self.test_dir = tempfile.mkdtemp()
        shutil.copy2(
            os.path.join(TEST_DIR, "Anforderungsübersicht_template.xlsx"), self.test_dir
        )

        data = os.path.join(TEST_DIR, "Anforderungsübersicht_data.xlsx.cdbxml.zip")
        template = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        ExcelReportGenerator(template, data).generate()

    def test_correct_date(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]
        report_date = "14.10.2022"
        expected_report_datetime = datetime.strptime(report_date, "%d.%m.%Y")

        datetime_in_sheet = sheet["I1"].value

        self.assertTrue(sheet["I1"].is_date)
        self.assertEqual(datetime_in_sheet, expected_report_datetime)
        workbook.close()

    def test_table_row_values(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]
        # list with filled at B, C, D, E, H (index)
        expected_values = (
            "idx0",
            "idx1",
            " 10",
            "Anf. mit Akzeptanz Kriteria",
            "RT000000090",
            "Anforderung",
            "idx6",
            "idx7",
            0,
            "idx8",
            "idx9",
            "idx10",
        )
        filled_columns_in_row = (2, 3, 4, 5, 8)
        for column in filled_columns_in_row:
            cell = get_column_letter(column) + "29"
            cell_value = sheet[cell].value
            expected_cell_value = expected_values[column]
            self.assertEqual(cell_value, expected_cell_value)  # add assertion here
        workbook.close()

    def test_worksheet_color(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]
        expected_headline_color_rgb = (0, 128, 197)

        headline_color = sheet["B1"].fill.start_color.index
        headline_color_hex = "#" + headline_color[2:]
        headline_color_rgb = ImageColor.getrgb(headline_color_hex)

        self.assertEqual(expected_headline_color_rgb, headline_color_rgb)

        expected_body_color_rgb = (228, 241, 248)

        body_color = sheet["K26"].fill.start_color.index
        body_color_hex = "#" + body_color[2:]
        body_color_rgb = ImageColor.getrgb(body_color_hex)

        self.assertEqual(expected_body_color_rgb, body_color_rgb)
        workbook.close()

    def test_group_at(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]

        expected_group_at = (30, 31, 32)
        rows = (28, 29, 30, 31, 32, 33)

        for row in rows:
            if row in expected_group_at:
                self.assertEqual(sheet.row_dimensions[row].outline_level, 1)
            else:
                self.assertNotEqual(sheet.row_dimensions[row].outline_level, 1)
        workbook.close()

    def test_color_grouping(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]

        expected_color_group0 = (230, 230, 255)
        expected_color_group1 = (210, 215, 243)  # (230, 230, 255) - (20, 20, 12)

        color_group0_hex = sheet["B28"].fill.start_color.index
        color_group0_hex_code = "#" + color_group0_hex[2:]

        color_group0_rgb = ImageColor.getrgb(color_group0_hex_code)

        # test next group color is correct
        self.assertEqual(color_group0_rgb, expected_color_group0)

        color_group1_hex = sheet["B30"].fill.start_color.index
        color_group1_hex_code = "#" + color_group1_hex[2:]

        color_group1_rgb = ImageColor.getrgb(color_group1_hex_code)

        self.assertEqual(color_group1_rgb, expected_color_group1)
        workbook.close()

    def test_grouping_indent(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]

        expected_indent_group0 = 0
        expected_indent_group1 = 1

        group0_indent = sheet["B28"].alignment.indent
        group1_indent = sheet["B30"].alignment.indent

        self.assertEqual(group0_indent, expected_indent_group0)
        self.assertEqual(group1_indent, expected_indent_group1)
        workbook.close()

    def test_data_validation(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]
        table_validation = sheet.data_validations.dataValidation[0]

        # test if validation exists
        self.assertTrue(table_validation)

        validation_formula = table_validation.formula1

        self.assertEqual(validation_formula, "Bewertungen")

        sheet = workbook["Tabelle3"]
        date_validation_table = sheet.tables["Tabelle13"]

        self.assertEqual(sheet["A1"].value, validation_formula)
        self.assertEqual(date_validation_table.ref, "A1:A6")
        workbook.close()

    def test_cells_has_hyperlinks_RT(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]

        expected_cell_type = "hyperlink"

        cell = sheet["D28"]
        cell_type = cell.hyperlink.tagname
        cell_text = cell.value

        self.assertEqual(cell_type, expected_cell_type)
        self.assertEqual(cell_text, "RT000000080")
        workbook.close()

    def test_cells_has_hyperlinks_AT(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]

        expected_cell_type = "hyperlink"

        cell = sheet["D30"]
        cell_type = cell.hyperlink.tagname
        cell_text = cell.value

        self.assertEqual(cell_type, expected_cell_type)
        self.assertEqual(cell_text, "AT0000014")
        workbook.close()

    def test_table_style(self):
        excel_file = os.path.join(self.test_dir, "Anforderungsübersicht_template.xlsx")
        workbook = xl.load_workbook(excel_file)
        sheet = workbook["Tabelle1"]

        expected_table_style = "cdb_xml_tab"

        sheet_table = sheet.tables["Tabelle6"]
        table_style = sheet_table.tableStyleInfo.name

        self.assertEqual(expected_table_style, table_style)
        workbook.close()

    def tearDown(self):
        super(GeneratedExcelTestCase, self).tearDown()
        shutil.rmtree(self.test_dir)
