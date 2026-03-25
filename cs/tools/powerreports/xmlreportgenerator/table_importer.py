import logging
import re

from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from cs.tools.powerreports.xmlreportgenerator import cdb_image_tools, tools
from cs.tools.powerreports.xmlreportgenerator.handle_xml import XMLFile
from cs.tools.powerreports.xmlreportgenerator.multi_export_tables import MultiExports
from cs.tools.powerreports.xmlreportgenerator.table_groups import GroupCreator
from cs.tools.powerreports.xmlreportgenerator.tools import (
    letter_num_seperated as separate,
)
from cs.tools.powerreports.xmlreportgenerator.worksheet_tools import WorksheetTools

LOG = logging.getLogger(__name__)


class TableImporter(object):
    """
    Load xml of table from table xml file
    Get table information (size and header) of table from worksheet
    Change/Update table
    Use table header and cdbxml data to insert data in table

    :param worksheet
    :param string table_file_path: path of xml file for table
    :param table: table inside worksheet
    :param dict xml_data: cdbxml data
    :param string xml_data_file: Name of cdbxml-file, because images have cdbxml-file
                                 + image name as file name
    """

    def __init__(
        self, worksheet, table_labels, table, xml_data, table_file_path, xml_data_file
    ):
        self.worksheet = worksheet
        self.table_labels = table_labels
        self.table = table
        self.xml_data = xml_data
        self.xml_data_file = xml_data_file

        self._table_columns_data = XMLFile(table_file_path).get_in_xml_with_keylist(
            ["table", "tableColumns", "tableColumn"]
        )
        self._table_headers_with_provider = {}
        self._table_size_start_end = None
        self._table_data = {}
        self._table_with_data_length = None

        self._table_start, self._table_end = tools.get_start_end_of_table(
            self.table_size_start_end
        )

        self.update_table()

    @property
    def table_size_start_end(self):
        if not self._table_size_start_end:
            table_size = self.table.ref
            self._table_size_start_end = re.split(":", table_size)
        return self._table_size_start_end

    @property
    def table_headers_with_provider(self):
        """
        Return information (size and table headers)
        tables header give information, which data of given provider to insert in column
        current table_size (work with start_cell and end_cell later)

        :LOGGING if a table (column of table) has no xml ref, but column is in an table of type xml

        :return dict table_headers_with_provider
        :return list table_size_start_and_end
        """
        if not self._table_headers_with_provider:
            table_start, _ = tools.get_start_end_of_table(self.table_size_start_end)
            table_header_column = column_index_from_string(table_start[0])

            for table_column_parameter in self.table.tableColumns:
                if table_column_parameter.xmlColumnPr is not None:
                    provider_path = table_column_parameter.xmlColumnPr.xpath
                    provider_split_path = provider_path.split("/")
                    table_provider = provider_split_path[2]
                    # {'header': column in worksheet, ..., 'provider': provider
                    # {'chapter': 2, 'desc_long_de': 3}
                    self._table_headers_with_provider["provider"] = table_provider
                    self._table_headers_with_provider[
                        table_column_parameter.uniqueName
                    ] = table_header_column
                    table_header_column += 1
                else:
                    table_header_column += 1

        return self._table_headers_with_provider

    @property
    def table_data_info(self):
        """
        Returns table data and length of table (if it would have data)
        """
        if not self._table_data and not self._table_with_data_length:
            provider = self.table_headers_with_provider["provider"]
            if provider in self.xml_data.keys():
                xml_data = self.xml_data[provider]
                xml_data = xml_data["List"]
                if self.table.headerRowCount == 1:
                    first_row = int(self._table_start[1]) + 1
                else:
                    first_row = int(self._table_start[1])
                if isinstance(xml_data, dict):
                    xml_data = [xml_data]

                for idx, data in enumerate(xml_data):
                    for table_column_ref, cell_value in data.items():
                        if table_column_ref in self.table_headers_with_provider:
                            cell_column = self.table_headers_with_provider[
                                table_column_ref
                            ]
                            cell_type = self.get_cell_type(table_column_ref)
                            cell_row = str(first_row + idx)
                            cell_address = get_column_letter(cell_column) + cell_row
                            self._table_data[cell_address] = [cell_value, cell_type]

                self._table_with_data_length = len(xml_data)
            else:
                self._table_with_data_length = 0

        return self._table_data, self._table_with_data_length

    def insert_in_table(self):
        """
        Insert values (from table_data) inside each cell
        """
        if self.table_data_info[0]:
            worksheet_images = tools.sheet_images(self.worksheet)
            cdb_temp_images = {}
            if self.table.headerRowCount == 1:
                first_row = int(self._table_start[1]) + 1
            else:
                first_row = int(self._table_start[1])
            init_row_height = self.worksheet.row_dimensions[first_row].height

            for image in worksheet_images:
                if "cdb_image" in image.anchor.pic.nvPicPr.cNvPr.name:
                    cdb_template_column = tools.anchor_from(image)
                    cdb_template_column = cdb_template_column.col + 1
                    cdb_template_column = get_column_letter(cdb_template_column)
                    cdb_temp_images[cdb_template_column] = image
                    image.anchor.pic.nvPicPr.cNvPr.hidden = "1"

            for table_cell_address, table_cell_data in self.table_data_info[0].items():
                table_cell_value = table_cell_data[0]

                if "cdb://image/" in table_cell_value:
                    temp_image = cdb_temp_images[table_cell_address[0]]
                    self.add_cdb_image(
                        temp_image,
                        table_cell_value,
                        table_cell_address,
                        init_row_height,
                    )
                elif (
                    " cdb:texttodisplay:" in table_cell_value
                    and self.worksheet.title != "cdb_template"
                ):
                    WorksheetTools(self.worksheet).create_cell_hyperlink(
                        table_cell_value, table_cell_address
                    )
                else:
                    try:
                        WorksheetTools(self.worksheet).set_cell_value(
                            table_cell_value, table_cell_data[1], table_cell_address
                        )
                    except ValueError:
                        attribute = ""
                        for attr, column in self.table_headers_with_provider.items():
                            if column == column_index_from_string(
                                separate(table_cell_address)[0]
                            ):
                                attribute = attr
                        raise ValueError(
                            "Expected attribute '%s' to be '%s' but value given '%s'. XSD Schema type definition '%s' of the attribute does not match."  # noqa
                            % (
                                attribute,
                                table_cell_data[1],
                                table_cell_value,
                                table_cell_data[1],
                            )
                        )

    def add_cdb_image(
        self, temp_image, table_cell_value, table_cell_address, row_height
    ):
        """
        Adding cdb_image to cell

        :param dict temp_image: template image of cell to replace with cdb_image of cml data
        :param string table_cell_value: value of cell where image has to be added
        :param string table_cell_address: address (column as letter and row as number) of cell
        :param int row_height: initial row height of first row, to have a default row_height

        Get temp_image with column of table_cell_address
        Get image settings (descriptions) from alternative text of temp_image

        Adding image using twoCellAnchor (edited as oneCell) like in drawing.xml of template file
        """
        (
            keep_width,
            keep_height,
            min_border_height,
            center_width,
            center_height,
        ) = cdb_image_tools.get_image_attribute_values(
            temp_image.anchor.pic.nvPicPr.cNvPr.descr
        )

        image = cdb_image_tools.get_image_to_add(
            table_cell_value, self.xml_data_file[:-7]
        )

        new_row_height = cdb_image_tools.change_image_get_row_height(
            image, temp_image, keep_height, keep_width, min_border_height, row_height
        )

        self.worksheet.row_dimensions[
            int(table_cell_address[1])
        ].height = new_row_height

        from_col, to_col, from_row, to_row = cdb_image_tools.get_anchor_attr(
            image, temp_image, center_width, center_height, new_row_height
        )

        marker_from = AnchorMarker(
            col=column_index_from_string(table_cell_address[0]) - 1,
            colOff=from_col,
            row=int(table_cell_address[1]) - 1,
            rowOff=from_row,
        )
        marker_to = AnchorMarker(
            col=column_index_from_string(table_cell_address[0]) - 1,
            colOff=to_col,
            row=int(table_cell_address[1]) - 1,
            rowOff=to_row,
        )
        image.anchor = TwoCellAnchor(editAs="oneCell", _from=marker_from, to=marker_to)

        self.worksheet.add_image(image)

    def get_cell_type(self, cell_ref):
        cell_type = ""

        for table_ref in self._table_columns_data:
            if table_ref["uniqueName"] == cell_ref:
                if "calculatedColumnFormula" in table_ref:
                    cell_type = "string"
                else:
                    cell_type = table_ref["xmlColumnPr"]["xmlDataType"]
        return cell_type

    def update_table(self):
        """
        Update table with new size and styles of first to all new rows

        Additional: extend data_validation for all cells in column
        Additional: Create hierarchy/groups
        """
        for label in self.table_labels:
            if label == self.table_headers_with_provider["provider"] + "_Labels":
                self.change_col_labels()
        if self.table_data_info[0]:
            new_table_size = tools.get_new_table_size(
                self._table_start, self._table_end, self.table_data_info[1]
            )
            new_table_size_start_end = re.split(":", new_table_size)
            max_rows = self.worksheet.max_row
            if max_rows != int(self._table_end[1]):
                moves_range = (
                    self._table_start[0]
                    + str(int(self._table_end[1]) + 1)
                    + ":"
                    + self._table_end[0]
                    + str(max_rows)
                )
                # self._table.headerRowCount is 1 or 0
                # to there is no other check needed if table has header or not
                table_length = (
                    int(self._table_end[1])
                    - int(self._table_start[1])
                    + 1
                    - self.table.headerRowCount
                )
                if table_length == 1 or (
                    self.table.totalsRowShown is not False and table_length == 2
                ):
                    moved_rows = self.table_data_info[1] - 1
                else:
                    moved_rows = self.table_data_info[1] - table_length
                self.worksheet.move_range(moves_range, rows=moved_rows, cols=0)
                self.change_table_ref_for_tables_below(
                    moves_range, moved_rows, self.table.displayName
                )
            if self.table.totalsRowShown is not False:
                if self.table.totalsRowCount is None:
                    self.table.totalsRowShown = False
                else:
                    new_table_size = tools.get_new_table_size(
                        self._table_start, self._table_end, self.table_data_info[1] + 1
                    )
                    total_row_pos = (
                        self._table_start[0]
                        + self._table_end[1]
                        + ":"
                        + "".join(self._table_end)
                    )
                    self.worksheet.move_range(
                        total_row_pos, rows=self.table_data_info[1] - 1, cols=0
                    )
            self.table.ref = new_table_size

            format_map = self.get_formatting_map()

            if format_map:
                self.worksheet.conditional_formatting = ConditionalFormattingList()
                for cells, rule in format_map.items():
                    self.worksheet.conditional_formatting.add(cells, rule)

            WorksheetTools(self.worksheet).copy_style_from_first_row(
                self.table, new_table_size_start_end
            )

            table_data_validation = self.worksheet.data_validations.dataValidation
            if table_data_validation:
                WorksheetTools(self.worksheet).extend_validations_by_table_length(
                    table_data_validation,
                    self._table_start,
                    self._table_end,
                    self.table_data_info[1],
                )

            self.extend_calculated_column_formula(new_table_size)

            for table_header in self.table_headers_with_provider:
                if "cdbxml_level" in table_header or "level:" in table_header:
                    cdbxml_level_position = self.table_headers_with_provider[
                        "cdbxml_level"
                    ]
                    GroupCreator(
                        self.worksheet,
                        cdbxml_level_position,
                        self.table_data_info[0],
                        new_table_size_start_end,
                    ).create_groups_in_table()

    def extend_calculated_column_formula(self, new_table_size):
        """
        Extends the formula(s) of the calculatedColumnFormula (table columns with formula) to every row

        :param str new_table_size: new (extended) table size e.g. "A2:T7"

        OpenPyXL does not automatically extend the calculatedColumnFormula in the manuel extended table
        With the new table size, the formula of a column have to be added to all cells of the column below

        The first and last row has to be calculated using new_table_size
        while consideration two factors (total row and header row) to prevent broken Excel files

        """
        for col_idx, column in enumerate(self.table.tableColumns):
            if column.calculatedColumnFormula:
                new_start, new_end = tools.get_start_end_of_table(re.split(":", new_table_size))
                # start row of table (if header is shown first row of table is header row)
                start_row = int(new_start[1]) if self.table.headerRowCount == 1 else int(new_start[1]) - 1
                # table with total rows has length +1 -> the column formula in total row cause error in Excel
                end_row = int(new_end[1]) if self.table.totalsRowShown is False else int(new_end[1]) - 1

                for row_idx in range(start_row + 1, end_row + 1):
                    table_column_start = column_index_from_string(self._table_start[0])
                    formula_cell = get_column_letter(col_idx + table_column_start) + str(row_idx)
                    self.worksheet[formula_cell].value = '=' + column.calculatedColumnFormula.attr_text

    def get_formatting_map(self):
        format_map = {}

        for worksheet_format in self.worksheet.conditional_formatting:
            # semicolon separated formatting cells/range for rule
            for formatting_cell in worksheet_format.cells:
                # check if formatting has range or is one cell only
                if ":" in str(formatting_cell):
                    start_cell, end_cell = str(formatting_cell).split(":")
                    start_col_row = separate(start_cell)
                    end_col_row = separate(end_cell)
                else:
                    start_cell = str(formatting_cell)
                    end_cell = start_cell
                    start_col_row = separate(start_cell)
                    end_col_row = separate(end_cell)

                # check if formatting is in table
                if start_col_row[1] in range(int(self._table_start[1]), int(self._table_end[1]) + 1) and \
                        column_index_from_string(start_col_row[0]) in \
                        range(column_index_from_string(self._table_start[0]),
                              column_index_from_string(self._table_end[0]) + 1,):
                    table_range = (start_cell + ":"
                                   + end_col_row[0] + str(self.table_data_info[1] - 1 + int(end_col_row[1])))
                else:
                    table_range = str(formatting_cell)
                format_map[table_range] = worksheet_format.cfRule[0]

        return format_map

    def change_col_labels(self):
        col_num = 0
        for table_column_parameter in self.table.tableColumns:
            curr_col = column_index_from_string(self._table_start[0]) + col_num
            get_label = self.worksheet.cell(
                row=int(self._table_start[1]) - 1, column=curr_col
            ).value
            selected_cell = get_column_letter(curr_col) + self._table_start[1]
            if (
                get_label is not None
                or "level:" not in self.worksheet[selected_cell].value
                or "cdbxml_level" not in self.worksheet[selected_cell].value
            ):
                table_column_parameter.name = str(get_label)
                self.worksheet[selected_cell] = str(get_label)
            col_num += 1

    def change_table_ref_for_tables_below(self, moves_range, moved_rows, table_name):
        moves_range_start, _ = moves_range.split(":")
        if "cdb_template" in self.worksheet.title:
            MultiExports(self.worksheet).move_row_height(moved_rows)
        for sheet_table in self.worksheet.tables.values():
            table_start, table_end = sheet_table.ref.split(":")
            if separate(table_start)[1] >= separate(moves_range_start)[1]:
                table_start_col, table_start_row = separate(table_start)
                table_end_col, table_end_row = separate(table_end)
                if column_index_from_string(table_end_col) > column_index_from_string(
                    self._table_end[0]
                ):
                    raise Exception(  # pylint: disable=W0719
                        "%s is wider than %s above. "
                        "The number of columns of the table below has to be equal or less "
                        "than the number of columns above."
                        % (sheet_table.displayName, table_name)
                    )
                table_start_row = str(table_start_row + moved_rows)
                table_end_row = str(table_end_row + moved_rows)
                sheet_table.ref = (
                    table_start_col
                    + table_start_row
                    + ":"
                    + table_end_col
                    + table_end_row
                )
