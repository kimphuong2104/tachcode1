import logging

from openpyxl.packaging.custom import StringProperty
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from cs.tools.powerreports.xmlreportgenerator import cdb_image_tools, tools
from cs.tools.powerreports.xmlreportgenerator.handle_xml import XMLFile
from cs.tools.powerreports.xmlreportgenerator.tools import (
    letter_num_seperated as separate,
)
from cs.tools.powerreports.xmlreportgenerator.worksheet_tools import WorksheetTools

LOG = logging.getLogger(__name__)


class SingleCellsImporter(object):
    """
    Load xml of single cells from single cells xml file
    Get data of single cells from single cells xml
    Insert data using provider of single cells from cdbxml

    :param workbook: excel workbook
    :param dict custom_props: dict with {provider_xsd_name: [attributes]} of single cell
                              to save in custom props
    :param worksheet
    :param string single_cells_file_path: path of xml file for single cells
    :param dict xml_data: cdbxml data
    :param string xml_data_file: Name of cdbxml-file, because images have cdbxml-file
                                 + image name as file name
    """

    def __init__(
        self,
        workbook,
        custom_props,
        worksheet,
        single_cells_file_path,
        xml_data,
        multi_export,
        xml_data_file,
    ):
        self.workbook = workbook
        self.custom_props = custom_props
        self.worksheet = worksheet
        self.xml_data = xml_data
        self.multi_export = multi_export
        self.xml_data_file = xml_data_file

        self._single_cell_xml = XMLFile(single_cells_file_path)
        self._single_cells_data = {}

    @property
    def single_cells_data(self):
        """
        Get Data from singleCells xml of Excel
        Data are cell_address with entry values to assgin cdbxml data to right single cell

        cell_address: [fieldname, provider, type]
        'I1': ['cdbxml_report_date', 'Arguments', 'date']

        :return dict single_cells_entry_data key: cell_address values: entry data
        """
        if not self._single_cells_data:
            single_cell_xml_data = self._single_cell_xml.get_in_xml_with_keylist(
                ["singleXmlCells", "singleXmlCell"]
            )

            for single_cell_dict in single_cell_xml_data:
                cell_address = single_cell_dict["r"]

                if self.multi_export:
                    cel_row = separate(cell_address)[1] + self.multi_export[0]
                    cell_column = column_index_from_string(separate(cell_address)[0])
                    cell_column = get_column_letter(cell_column + self.multi_export[1])
                    cell_address = cell_column + str(cel_row)

                single_cell = single_cell_dict["xmlCellPr"]["xmlPr"]
                cell_path = single_cell["xpath"]
                cell_type = single_cell["xmlDataType"]

                fieldname, provider = tools.get_fieldname_and_provider_from_path(
                    cell_path
                )
                self._single_cells_data[cell_address] = {
                    "fieldname": fieldname,
                    "provider": provider,
                    "type": cell_type,
                }
        return self._single_cells_data

    def insert_in_single_cells(self):
        """
        Insert values (if in cdbxml) in single cells of the sheet

        Each item of data is one cell
        Getting address of cell and the associated entry data for the address from single_cells_data

        Check if provider of data are in cdbxml
        Check if fieldname (from single_cell_data) is in provider

        Get value using fieldname in provider from cdbxml
        Change type of single cell value and insert to worksheet

        :LOGGING if in single cells has missing data
        :LOGGING if single cells should have data of a provider, which is not in cdbxml
        :return list table_labels:  if provider is _labels for tables, provider name is saved in list
        """
        table_labels = []
        self.set_default_props()
        for (
            single_cell_address,
            single_cell_entry_data,
        ) in self.single_cells_data.items():
            fieldname = single_cell_entry_data["fieldname"]
            single_cell_value_type = single_cell_entry_data["type"]
            provider = single_cell_entry_data["provider"]

            if provider not in self.xml_data.keys():
                LOG.warning("Provider %s not in data.", provider)
            else:
                single_cells_from_provider = self.xml_data[provider]
                if fieldname in single_cells_from_provider:
                    single_cell_value = single_cells_from_provider[fieldname]

                    self.set_custom_props(provider, fieldname, single_cell_value)

                    if "cdb://image/" in single_cell_value:
                        cdb_temp_images = self.get_cdb_temp_images()
                        temp_image = cdb_temp_images[single_cell_address[0]]
                        self.add_cdb_image(
                            temp_image, single_cell_value, single_cell_address
                        )
                    elif (
                        " cdb:texttodisplay:" in single_cell_value
                        and self.worksheet.title != "cdb_template"
                    ):
                        WorksheetTools(self.worksheet).create_cell_hyperlink(
                            single_cell_value, single_cell_address
                        )
                    else:
                        try:
                            WorksheetTools(self.worksheet).set_cell_value(
                                single_cell_value,
                                single_cell_value_type,
                                single_cell_address,
                            )
                        except ValueError:
                            raise ValueError(
                                "Expected attribute '%s' to be '%s' but value given '%s'. XSD Schema type definition '%s' of the attribute does not match."  # noqa
                                % (
                                    fieldname,
                                    single_cell_value_type,
                                    single_cell_value,
                                    single_cell_value_type,
                                )
                            )
                    if "_Labels" in provider and provider not in table_labels:
                        table_labels.append(provider)
                else:
                    LOG.warning(
                        "Data reference %s from %s is missing in data.",
                        fieldname,
                        provider,
                    )
        return table_labels

    def set_default_props(self):
        if (
            "Arguments" in self.custom_props.keys()
            and "Arguments" in self.xml_data.keys()
        ):
            for default_attr in self.custom_props["Arguments"]:
                attr_provider = self.xml_data["Arguments"]
                prop_name = "Arguments." + default_attr
                self.add_custom_prop(prop_name, attr_provider[default_attr])

    def set_custom_props(self, provider, fieldname, single_cell_value):
        if (
            provider in self.custom_props.keys()
            and fieldname in self.custom_props[provider]
        ):
            prop_name = provider + "." + fieldname
            self.add_custom_prop(prop_name, single_cell_value)

    def add_custom_prop(self, prop_name, prop_value):
        if prop_name not in [
            prop.name for prop in self.workbook.custom_doc_props.props
        ]:
            self.workbook.custom_doc_props.append(
                StringProperty(name=prop_name, value=prop_value)
            )

    def get_cdb_temp_images(self):
        worksheet_images = tools.sheet_images(self.worksheet)
        cdb_temp_images = {}
        for image in worksheet_images:
            if (
                type(image.anchor) is not str
                and "cdb_image" in image.anchor.pic.nvPicPr.cNvPr.name
            ):
                cdb_template_column = tools.anchor_from(image)
                cdb_template_column = cdb_template_column.col + 1
                cdb_template_column = get_column_letter(cdb_template_column)
                cdb_temp_images[cdb_template_column] = image
                image.anchor.pic.nvPicPr.cNvPr.hidden = "1"
        return cdb_temp_images

    def add_cdb_image(self, temp_image, single_cell_value, single_cell_address):
        keep_width, keep_height, _, _, _ = cdb_image_tools.get_image_attribute_values(
            temp_image.anchor.pic.nvPicPr.cNvPr.descr
        )

        image = cdb_image_tools.get_image_to_add(
            single_cell_value, self.xml_data_file[:-7]
        )

        temp_image_height = temp_image.height
        temp_image_width = temp_image.width

        image.width, image.height = cdb_image_tools.get_new_image_h_w(
            image, temp_image_height, temp_image_width, keep_height, keep_width
        )
        image.anchor = single_cell_address
        self.worksheet.add_image(image)
