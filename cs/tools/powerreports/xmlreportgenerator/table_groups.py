import re
from copy import copy

import openpyxl as xl
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from cs.tools.powerreports.xmlreportgenerator import tools


class GroupCreator(object):
    """
    Search for groups
    Create dict with groups and level of group
    Create group in table in worksheet
    Style table (the groups) with color and intent depends on settings

    :param worksheet
    :param int group_level_at: column where group information is
    :param dict table_data: table data also contains cdbxml_level value 0, 1, 2 etc
    :param list table_size_start_to_end: list with 2 elements, start and end cells of table
    """

    def __init__(self, worksheet, group_level_at, table_data, table_size_start_to_end):
        self.worksheet = worksheet
        self.table_data = table_data
        self.table_size_start_to_end = table_size_start_to_end
        self.group_level_at = group_level_at

        self._rows_and_level = {}
        self.style_group()

    @property
    def rows_and_level(self):
        """
        Creates dict with row and level of row

        :return dict rows_and_level: key = row, value = level
        """
        if not self._rows_and_level:  # pylint: disable=R1702
            group_level_with_rows = self.get_groups_with_level()
            if group_level_with_rows:
                for level, rows_by_level in group_level_with_rows.items():
                    # level 0 marks start of first group level 1 - can be ignored -
                    # later automatic added by openpyxl
                    # 0: ['27', '40', '49'],
                    # 1: ['28', '29' ... '41', '42', '43', '44', '45', '46', '47', '48' ...]
                    if level != 0:
                        for row in rows_by_level:
                            # added row x to level
                            self._rows_and_level[int(row)] = level

        return self._rows_and_level

    def create_groups_in_table(self):
        """
        Creates groups using row_dimensions from openpyxl with ranges_for_leveling
        """
        if self.worksheet.title != "cdb_template":
            for row, row_level in self.rows_and_level.items():
                self.worksheet.row_dimensions.group(row, row, outline_level=row_level)

    def get_groups_attributes(self):
        """
        get attribute color and indent of group from cell with address (table_start , _group_level_at)

        :return string group_color: colorize(195,195,195),(-25,-25,-25) or None
        :return string group_indent: indent(1) or None
        """
        table_start, _ = tools.get_start_end_of_table(self.table_size_start_to_end)
        group_color = None
        group_indent = None

        cdbxml_level_values = self.worksheet.cell(
            int(table_start[1]), self.group_level_at
        ).value
        if "level:" in cdbxml_level_values:
            cdbxml_level_styles = re.split(":", cdbxml_level_values)
            for style in cdbxml_level_styles:
                if "indent" in style:
                    group_indent = style
                if "colorize" in style:
                    group_color = style

        return group_color, group_indent

    def get_groups_with_level(self):
        """
        Return created dict for with groups level and rows of level

        :return dict group_level_with_rows
        """
        group_level_with_rows = {}
        for cell_location, cell_entry_data in self.table_data.items():
            cell_value = cell_entry_data[0]
            group_level_column = get_column_letter(self.group_level_at)
            _, group_level_row = tools.letter_num_seperated(cell_location)
            # only add group level if cell address (column) is group_level_at_col
            if group_level_column in cell_location:
                try:
                    group_level_depth = int(cell_value)
                except ValueError:
                    raise ValueError(
                        "Attribute 'cdbxml_level' has value %s and is not type integer."
                        % cell_value
                    )
                # if group level not in dict - create new key of group_level_depth
                if group_level_depth not in group_level_with_rows:
                    group_level_with_rows[group_level_depth] = [group_level_row]
                # if group row not in group level - append to level
                if group_level_row not in group_level_with_rows[group_level_depth]:
                    group_level_with_rows[group_level_depth].append(group_level_row)

        # reverse dict to start with last level x to first level 0
        return tools.get_dict_reversed(group_level_with_rows)

    def style_group(self):
        """
        Style rows and columns of groups with given colors and indent based of level

        Sometimes first column of table mustn't be A (1)
        In case start_column is B (2) and column_with_indent = 2
        Column C (3) has indent
        """
        table_start, table_end = tools.get_start_end_of_table(
            self.table_size_start_to_end
        )
        group_color, group_indent = self.get_groups_attributes()
        start_row = int(table_start[1]) + 1
        end_row = int(table_end[1]) + 1
        start_column = column_index_from_string(table_start[0])
        end_column = column_index_from_string(table_end[0])
        base_color_rgb, color_offset_rgb = None, None

        for row in range(start_row, end_row):
            for column in range(start_column, end_column):
                # colorize rows and columns with base_color
                if group_color is not None:
                    base_color_rgb, color_offset_rgb = tools.get_base_offset_color(
                        group_color
                    )
                    self.fill_cell_background_color(
                        row, column, tools.get_hex_of_color_rgb(base_color_rgb)
                    )

                # for each level colorize and indent rows and columns
                for r, row_level in self.rows_and_level.items():
                    if row in range(r, r + 1):
                        if group_indent is not None:
                            column_with_indent = self.get_column_with_indent(
                                start_column, group_indent
                            )
                            column_alignment = copy(
                                self.worksheet.cell(row, column_with_indent).alignment
                            )
                            column_alignment.indent = row_level
                            self.worksheet.cell(
                                row, column_with_indent
                            ).alignment = copy(column_alignment)

                        if base_color_rgb is not None and color_offset_rgb is not None:
                            color_offset_rgb_level = tuple(
                                row_level * rgb for rgb in color_offset_rgb
                            )
                            color_rgb = tools.calculate_color(
                                base_color_rgb, color_offset_rgb_level
                            )
                            self.fill_cell_background_color(
                                row, column, tools.get_hex_of_color_rgb(color_rgb)
                            )

    @staticmethod
    def get_column_with_indent(start_column, group_indent):
        column_with_indent = int(re.search(r"\((.*?)\)", group_indent).group(1))
        if column_with_indent == 1:
            column = start_column
        else:
            column = start_column + column_with_indent - 1
        return column

    def fill_cell_background_color(self, row, column, color_hex):
        self.worksheet.cell(row, column).fill = xl.styles.PatternFill(
            fill_type="solid", start_color=color_hex
        )
