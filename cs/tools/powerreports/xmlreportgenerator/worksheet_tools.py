from copy import copy, deepcopy

import openpyxl as xl
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from cs.tools.powerreports.xmlreportgenerator import tools


class WorksheetTools(object):
    """
    Tools that use worksheet and are used by TableImporter and SingleCellsImporter
    :param worksheet: current worksheet
    """

    def __init__(self, worksheet):
        self.worksheet = worksheet

    def create_cell_hyperlink(self, input_value, cell_address):
        """
        :param input_value: value of input to cell (has to be changed to hyperlink)
        :param cell_address: A3 B5 etc where hyperlink should be

        Creates a hyperlink for the cell
        Get url and text from value with string split at ' cdb:texttodisplay:' (powerreports/__init__.py)
        Change cell url, text and color
        """
        value = input_value.split(" cdb:texttodisplay:")
        url = value[0]
        text_to_display = value[1]
        self.worksheet[cell_address].value = text_to_display
        self.worksheet[cell_address].hyperlink = url
        self.worksheet[cell_address].font = xl.styles.Font(
            u="single", color=xl.styles.colors.BLUE
        )

    def set_cell_value(self, cell_value, cell_value_type, cell_address):
        """
        :param cell_value: raw value of cell
        :param cell_value_type: wanted type (date/dateTime, int, float or string) of cell value
        :param cell_address: destination cell to insert value

        Set value of cell in worksheet
        Change type of value beforehand into the known/wanted type
        """
        new_cell_value = cell_value
        if cell_value_type in ("date", "dateTime"):
            new_cell_value = tools.get_date_time(cell_value)
        elif cell_value_type == "integer":
            new_cell_value = int(cell_value)
        elif cell_value_type == "float":
            new_cell_value = float(cell_value)
        self.worksheet[cell_address].value = new_cell_value

    def extend_validations_by_table_length(
        self, table_data_validations, table_start, table_end, table_length
    ):
        """
        :param table_data_validations: list of all data validations in table
        :param table_start: start cell of table
        :param table_end: end cell of table
        :param table_length: length (in rows) of table

        Extend the data validation of column by length of table
        (is not extended automatically with the change of the table size)

        table can have multiple data validations
        using idx to extend every data validation in table
        """

        for validation_idx in range(len(table_data_validations)):
            table_data_validation = self.worksheet.data_validations.dataValidation[
                validation_idx
            ]
            table_column_with_validation = table_data_validation.sqref

            table_header_column_range = range(
                column_index_from_string(table_start[0]),
                column_index_from_string(table_end[0]) + 1,
            )

            data_validation_column, data_validation_row = tools.letter_num_seperated(
                str(table_column_with_validation)
            )

            data_validation_column_number = column_index_from_string(
                data_validation_column
            )

            if (
                data_validation_column_number in table_header_column_range
                and str(data_validation_row) == table_end[1]
            ):
                table_data_validation.sqref = (
                    str(table_column_with_validation)
                    + ":"
                    + data_validation_column
                    + str(int(data_validation_row) + table_length - 1)
                )

    def copy_style_from_first_row(self, table, table_start_to_end):
        """
        :param table: current table
        :param table_start_to_end: string with start-end of table like A18:K19

        Copy style of first row to every follow up row in table
        Style is for each column of first row different
        """
        table_start, _ = tools.get_start_end_of_table(table_start_to_end)
        if table.headerRowCount == 1:
            first_row = int(table_start[1]) + 1
        else:
            first_row = int(table_start[1])
        table_start[1] = str(first_row)
        range_table = table_start[0] + table_start[1] + ":" + table_start_to_end[1]

        for row in self.worksheet[range_table]:
            for cell in row:
                sheet_coordinate = self.worksheet[cell.coordinate]
                sheet_coordinate.alignment = copy(
                    self.worksheet.cell(first_row, cell.column).alignment
                )
                sheet_coordinate.font = copy(
                    self.worksheet.cell(first_row, cell.column).font
                )
                sheet_coordinate.fill = copy(
                    self.worksheet.cell(first_row, cell.column).fill
                )
                sheet_coordinate.border = copy(
                    self.worksheet.cell(first_row, cell.column).border
                )
                sheet_coordinate.number_format = copy(
                    self.worksheet.cell(first_row, cell.column).number_format
                )
                sheet_coordinate.protection = copy(
                    self.worksheet.cell(first_row, cell.column).protection
                )

    def copy_cdb_template_worksheet(self, workbook, key):
        cdb_template_copy = workbook.copy_worksheet(self.worksheet)
        for table in self.worksheet.tables.values():
            # use deepcopy, so original names wont be changed (happens with normal copy)
            # double table names not allowed in excel
            tab = deepcopy(table)
            # add _ to keep original name
            tab.name = table.name + "_" + str(key)
            cdb_template_copy.add_table(tab)
        return cdb_template_copy

    def get_hidden_rows_columns(self, offsets):
        hidden_rows_columns = {"Rows": [], "Table": [], "Columns": []}
        if len(self.worksheet.tables) != 0:
            for table in self.worksheet.tables.values():
                table_start, _ = table.ref.split(":")
                table_header_row = (
                    tools.letter_num_seperated(table_start)[1] + offsets[0]
                )
                hidden_rows_columns["Table"].append(table_header_row)
        if self.worksheet.max_row > self.worksheet.max_column:
            test = int(self.worksheet.max_row) + 2
        else:
            test = int(self.worksheet.max_column) + 2
        for idx in range(1, test):
            if self.worksheet.row_dimensions[idx].hidden:
                hidden_rows_columns["Rows"].append(idx + offsets[0])
            if self.worksheet.column_dimensions[get_column_letter(idx)].hidden:
                hidden_rows_columns["Columns"].append(idx + offsets[1])
        return hidden_rows_columns

    def return_multi_export_infos(self, multi_export_infos, key, export_start):
        table_start, table_end, table_style = None, None, None
        sheet_tables = {}
        row_sizes = []
        for row in range(
            int(export_start[1:]), self.worksheet.max_row + int(export_start[1:])
        ):
            size = self.worksheet.row_dimensions[row].height
            row_sizes.append(size)
        if len(self.worksheet.tables) == 1:
            for table in self.worksheet.tables.values():
                table_start, table_end = table.ref.split(":")
                table_style = table.tableStyleInfo

            multi_export_infos[key] = {
                "Start": table_start,
                "End": table_end,
                "Style": table_style,
                "Tables": None,
                "MaxRow": self.worksheet.max_row,
                "RowSizes": row_sizes,
            }

        elif len(self.worksheet.tables) > 1:
            idx = 0
            for table in self.worksheet.tables.values():
                table_start, table_end = table.ref.split(":")
                table_style = table.tableStyleInfo
                sheet_tables[idx] = {
                    "Start": table_start,
                    "End": table_end,
                    "Style": table_style,
                }
                idx += 1

            multi_export_infos[key] = {
                "Start": export_start,
                "End": table_end,
                "Style": "Tables",
                "Tables": sheet_tables,
                "MaxRow": self.worksheet.max_row - int(table_end[1:]),
                "RowSizes": row_sizes,
            }
        elif len(self.worksheet.tables) == 0:
            multi_export_infos[key] = {
                "Start": export_start,
                "End": get_column_letter(self.worksheet.max_column)
                + str(self.worksheet.max_row),
                "MaxRow": self.worksheet.max_row,
                "Style": None,
                "Tables": None,
                "RowSizes": row_sizes,
            }
        return multi_export_infos

    def get_merged_cells(self, workbook, multi_export_value):
        merged_cells_in_ws = []
        cdb_template_copy = workbook.copy_worksheet(self.worksheet)

        for cell_range in cdb_template_copy.merged_cells.ranges:
            cdb_template_copy.move_range(
                cell_range, rows=multi_export_value[0], cols=multi_export_value[1]
            )
            merged_cells_in_ws.append(str(cell_range))
        # add merged cells of cdb_template_copy to target
        workbook.remove(workbook[cdb_template_copy.title])
        return merged_cells_in_ws
