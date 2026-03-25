# -*- coding: utf-8 -*-
import logging
import os
import shutil

import openpyxl as xl
from openpyxl.packaging.custom import StringProperty
from openpyxl.utils.cell import column_index_from_string

from cs.tools.powerreports.reportserver.reportlib import getConfValue
from cs.tools.powerreports.xmlreportgenerator import tools
from cs.tools.powerreports.xmlreportgenerator.handle_xml import XMLFile
from cs.tools.powerreports.xmlreportgenerator.multi_export_tables import MultiExports
from cs.tools.powerreports.xmlreportgenerator.single_cell_importer import (
    SingleCellsImporter,
)
from cs.tools.powerreports.xmlreportgenerator.table_importer import TableImporter
from cs.tools.powerreports.xmlreportgenerator.tools import (
    letter_num_seperated as separate,
)
from cs.tools.powerreports.xmlreportgenerator.worksheet_tools import WorksheetTools

DEFAULT_CELL = "A1"
LOG = logging.getLogger(__name__)
DEBUG = getConfValue("REPORT_DEBUG", False)


class ExcelReportGenerator(object):
    def __init__(self, result_template_path, zip_path, custom_props=None):
        self.template_path = result_template_path
        self.xml_data_path = zip_path
        if not custom_props:
            custom_props = {}
        self.custom_props = custom_props

        self.data_basename = os.path.splitext(os.path.split(zip_path)[1])[0]

        self._cdbxml_temp = None
        self._template_dir = None
        self._file_map = None

    @property
    def file_map(self):
        if not self._file_map:
            self._file_map = self.get_file_map()
        return self._file_map

    @property
    def cdbxml_temp(self):
        if not self._cdbxml_temp:
            self._cdbxml_temp = tools.temporary_unzip_file(self.xml_data_path)
        return self._cdbxml_temp

    @property
    def template_dir(self):
        if not self._template_dir:
            self._template_dir = tools.temporary_unzip_file(self.template_path)
        return self._template_dir

    def generate(self, make_pdf=False):
        """
        Generate Excel

        workbook = Excel file
        worksheets = sheet(s) of Excel

        single cells = cells with xml ref, but are not in an table
        table = extra table in sheet of Excel

        Load workbook
        For every sheet in workbook method for single_cells an tables
        Save workbook under new name, otherwise template will be overwritten
        """
        try:
            _, extension = os.path.splitext(self.template_path)
            if extension == ".xlsm":
                workbook = xl.load_workbook(self.template_path, keep_vba=True)
                # only add VBAExecuted if excel is xslm
                workbook.custom_doc_props.append(
                    StringProperty(name="VBAExecuted", value="false")
                )
            else:
                workbook = xl.load_workbook(self.template_path, keep_vba=False)

            for worksheet_idx, worksheet in enumerate(workbook.worksheets):
                if worksheet.title != "cdb_template":
                    self.insert_to_sheet(
                        workbook, self.data_basename, worksheet, worksheet_idx
                    )
                else:
                    self.cdb_template_sheet(workbook, worksheet, worksheet_idx)

            workbook.save(filename=self.template_path)

            from cs.tools.powerreports.xmlreportgenerator import post_processing

            post_processing.vba_signature(self.template_dir, self.template_path)

            res = {
                "status": "OK",
                "xls": self.template_path,
                "pdf": None,
            }
            if make_pdf:
                res["pdf"] = self.create_pdf(self.template_path)
        except Exception as ex:  # pylint: disable=W0703
            LOG.error("Could not edit/save workbook: %s", ex, exc_info=True)
            res = {
                "status": ex,
                "xls": None,
                "pdf": None,
            }

        self.clean_up()

        return res

    def cdb_template_sheet(self, workbook, worksheet, worksheet_idx):
        target_sheet = workbook.worksheets[0]
        LOG.warning("Last row of cdb_template is %i.", worksheet.max_row)
        export_start = DEFAULT_CELL
        row_offset = 0
        multi_export_value = [0, 0]

        if "!$" in worksheet[DEFAULT_CELL].value:
            export_start = worksheet[DEFAULT_CELL].value.split("!$")[1]
            target_sheet = workbook[worksheet[DEFAULT_CELL].value.split("!$")[0]]
            if separate(export_start)[1] == 1:
                row_offset = -1
                multi_export_value = [
                    0,
                    column_index_from_string(separate(export_start)[0]) - 1,
                ]
            else:
                row_offset = separate(export_start)[1] - 1
                multi_export_value = [
                    row_offset,
                    column_index_from_string(separate(export_start)[0]) - 1,
                ]
            worksheet[DEFAULT_CELL].value = None
        file_keys = list(self.file_map.keys())
        # has to sort: if the files in cdbxml.zip are not by name sorted
        file_keys.sort(reverse=True)

        multi_export_infos = {}
        export_offsets = []

        hidden_rows_columns = WorksheetTools(worksheet).get_hidden_rows_columns(
            multi_export_value
        )

        merged_cells_in_ws = WorksheetTools(worksheet).get_merged_cells(
            workbook, multi_export_value
        )

        if export_start != DEFAULT_CELL:
            MultiExports(worksheet).move_cdb_template_input(multi_export_value)

        for key in file_keys:
            cdb_template = WorksheetTools(worksheet).copy_cdb_template_worksheet(
                workbook, key
            )

            self.insert_to_sheet(
                workbook,
                os.path.join(self.cdbxml_temp, self.file_map[key]),
                cdb_template,
                worksheet_idx,
                multi_export_value,
            )

            multi_export_infos = WorksheetTools(cdb_template).return_multi_export_infos(
                multi_export_infos, key, export_start
            )
            if row_offset >= 1:
                if multi_export_infos[key]["Style"] != "Tables":
                    amount_rows_to_insert = (
                        separate(multi_export_infos[key]["End"])[1] - row_offset
                    )
                else:
                    amount_rows_to_insert = (
                        separate(multi_export_infos[key]["End"])[1]
                        + multi_export_infos[key]["MaxRow"]
                        - row_offset
                    )
                export_offsets.append(amount_rows_to_insert)
                target_sheet.move_range(
                    export_start
                    + ":"
                    + multi_export_infos[key]["End"][0]
                    + str(sum(export_offsets) + amount_rows_to_insert),
                    rows=amount_rows_to_insert,
                    cols=0,
                )
            else:
                if multi_export_infos[key]["Style"] != "Tables":
                    amount_rows_to_insert = separate(multi_export_infos[key]["End"])[1]
                else:
                    amount_rows_to_insert = (
                        separate(multi_export_infos[key]["End"])[1]
                        + multi_export_infos[key]["MaxRow"]
                        - row_offset
                    )
                export_offsets.append(amount_rows_to_insert)
                target_sheet.move_range(
                    export_start
                    + ":"
                    + separate(multi_export_infos[key]["End"])[0]
                    + str(sum(export_offsets) + amount_rows_to_insert),
                    rows=amount_rows_to_insert,
                    cols=0,
                )
            if multi_export_infos[key]["Style"] == "Tables":
                MultiExports(target_sheet).copy_template_target(
                    cdb_template,
                    export_start,
                    multi_export_infos[key]["End"][0]
                    + str(
                        separate(multi_export_infos[key]["End"])[1]
                        + multi_export_infos[key]["MaxRow"]
                    ),
                )
            else:
                MultiExports(target_sheet).copy_template_target(
                    cdb_template, export_start, multi_export_infos[key]["End"]
                )

            workbook.remove(workbook[cdb_template.title])
        single_cell_size = separate(multi_export_infos[0]["Start"])[1]

        if row_offset == 0:
            export_offsets.append(0)
        elif row_offset == -1:
            export_offsets.append(0)
        else:
            single_cell_size = single_cell_size - (row_offset - 1)
            export_offsets.append(row_offset)

        MultiExports(target_sheet).add_formats_to_sheet(
            export_start,
            export_offsets,
            hidden_rows_columns,
            single_cell_size,
            multi_export_infos,
            merged_cells_in_ws,
        )
        worksheet.sheet_state = "veryHidden"

    def insert_to_sheet(
        self, workbook, filename, worksheet, worksheet_idx, multi_export=None
    ):
        cdbxml_file_path = os.path.join(self.cdbxml_temp, filename)
        if (
            os.path.exists(cdbxml_file_path)
            and "cdbxml" in os.path.splitext(cdbxml_file_path)[1]
        ):
            xml_data = XMLFile(cdbxml_file_path).get_single_key("Root")
            table_labels = self.fill_single_cells_in_sheet(
                workbook,
                worksheet,
                worksheet_idx,
                xml_data,
                multi_export,
                cdbxml_file_path if multi_export is None else None,
            )
            self.fill_tables_in_sheet(
                worksheet,
                xml_data,
                table_labels,
                cdbxml_file_path if multi_export is None else None,
            )

    def fill_single_cells_in_sheet(
        self,
        workbook,
        worksheet,
        worksheet_idx,
        xml_data,
        multi_export_values,
        cdbxml_file_path,
    ):
        """
        Insert data from cdbxml into single cells of worksheet

        :param  workbook: excel workbook
        :param  worksheet: current sheet of workbook
        :param  worksheet_idx: index of worksheet
        :param  xml_data: dict of cdbxml
        :param  multi_export_values: None if not multi-export; list [row, column] with offset
        :param  cdbxml_file_path: cdbxml temp name (used for cdb_image)

        Get from relation file of current sheet all contained relations (tables, singleCells, drawings etc)
        Get type of relation tableSingleCells and from that thr file path
        (path is ../tables/tableSingleCells1.xml)
        Real single_cells_file_path is xl/tables/tableSingleCells1.xml
        """
        sheet_file = "sheet" + str(worksheet_idx + 1) + ".xml"
        relation_file_path = os.path.join(
            self.template_dir, "xl", "worksheets", "_rels", sheet_file + ".rels"
        )
        table_labels = []
        if os.path.isfile(relation_file_path):
            relationships = XMLFile(relation_file_path).get_in_xml_with_keylist(
                ["Relationships", "Relationship"]
            )
            for relationship in relationships:
                relationship_type = relationship["Type"].split("/")[-1]
                if relationship_type == "tableSingleCells":
                    single_cells_filename = relationship["Target"].split("/")[-1]
                    single_cells_file_path = os.path.join(
                        self.template_dir, "xl", "tables", single_cells_filename
                    )
                    table_labels = SingleCellsImporter(
                        workbook,
                        self.custom_props,
                        worksheet,
                        single_cells_file_path,
                        xml_data,
                        multi_export_values,
                        cdbxml_file_path,
                    ).insert_in_single_cells()
        return table_labels

    def fill_tables_in_sheet(self, worksheet, xml_data, table_labels, cdbxml_file_path):
        """
        Insert data from cdbxml into an table inside worksheet (sheets of excel)
        Change Pivot Table to Refresh when Excel loads

        :param  worksheet: current sheet of workbook
        :param  table_labels: string if table has dynamic
        :param  xml_data: dict of cdbxml
        :param  cdbxml_file_path: cdbxml temp name (used for cdb_image)


        Get list of files in xl/tables
        For every table in sheet:
            If filename table1.xml or table2.xml (last 5 char removed) is table, get table xml file path
            If table (from xl/tables) name = table.name (from sheet), start table_import
            with table and xl/tables file
            Relationship file not needed for tables, because can check right table via names
        """
        table_dir_path = os.path.join(self.template_dir, "xl", "tables")
        if not os.path.exists(table_dir_path):
            raise Exception(  # pylint: disable=W0719
                "Report template has no xml scheme."
            )
        file_list = os.listdir(table_dir_path)
        for table in worksheet.tables.values():
            tab_name = table.name
            if table.tableType == "xml":
                if "cdb_template" in worksheet.title:
                    tab_name = table.name.split("_")[0]
                for table_filename in file_list:
                    if separate(os.path.splitext(table_filename)[0])[0] == "table":
                        table_file_path = os.path.join(table_dir_path, table_filename)
                        table_xml_data = XMLFile(table_file_path).get_single_key(
                            "table"
                        )
                        if table_xml_data["name"] == tab_name:
                            TableImporter(
                                worksheet,
                                table_labels,
                                table,
                                xml_data,
                                table_file_path,
                                cdbxml_file_path,
                            ).insert_in_table()
        tools.refresh_pivot(worksheet)

    def get_file_map(self):
        files = os.listdir(self.cdbxml_temp)
        if self.data_basename in files:
            files.remove(self.data_basename)

        for file_name in files[:]:
            if "cdbxml" not in os.path.splitext(file_name)[1]:
                files.remove(file_name)
        return dict(enumerate(files))

    def clean_up(self):
        if DEBUG:
            LOG.info("cdbxml directory: %s", self.cdbxml_temp)
            LOG.info("Template directory: %s", self.template_dir)
        else:
            try:
                shutil.rmtree(self.cdbxml_temp)
            except Exception as e:  # pylint: disable=W0703
                LOG.warning(
                    "Could not remove cdbxml directory: %s (%s).", self.cdbxml_temp, e
                )

            try:
                shutil.rmtree(self.template_dir)
            except Exception as e:  # pylint: disable=W0703
                LOG.warning(
                    "Could not remove template directory: %s (%s).",
                    self.template_dir,
                    e,
                )

    @staticmethod
    def create_pdf(fname):
        from cs.office.acsplugins.office import pdfconverter

        pdfconverter.convert(fname, cfgfile=None, znum=None, zidx=None)
        pdf_result = "%s.pdf" % os.path.splitext(fname)[0]
        return pdf_result
